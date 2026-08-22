"""Tendinha do Saber - Backend API (FastAPI + MongoDB)."""
from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import re
import asyncio
import time
import uuid
import json
import shutil
import logging
import secrets
import hashlib
import hmac
import mimetypes
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Dict, Any, Annotated

import bcrypt
import jwt
import pandas as pd
from io import BytesIO
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, UploadFile, File, Form, Query
from fastapi.responses import JSONResponse, FileResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr, ConfigDict

# ---------- Setup ----------
mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MIN = 60 * 24  # 24h
JWT_SECRET = os.environ["JWT_SECRET"]
LAMINATION_PRICE = float(os.environ.get("LAMINATION_PRICE", "2.00"))
SHIPPING_FLAT_RATE = float(os.environ.get("SHIPPING_FLAT_RATE", "4.90"))
BAG_PRICE = float(os.environ.get("BAG_PRICE", "0.10"))
STOCK_LOW_THRESHOLD = 5  # "pouco stock" no admin; ajuste fácil se necessário.
ISBNDB_API_KEY = os.environ.get("ISBNDB_API_KEY", "").strip()
ISBNDB_BASE_URL = "https://api2.isbndb.com"
ISBNDB_MIN_INTERVAL = 1.05
_ISBNDB_LOCK = asyncio.Lock()
_ISBNDB_LAST_REQUEST_AT = 0.0

# Private, non-public storage for MEGA voucher PDFs. NEVER place this under
# frontend/public or mount it as a FastAPI StaticFiles route — it must only
# ever be readable through the authenticated /admin/vouchers/{id}/pdf route.
VOUCHERS_DIR = ROOT_DIR / "private_storage" / "vouchers"
VOUCHERS_DIR.mkdir(parents=True, exist_ok=True)
VOUCHER_MAX_BYTES = 8 * 1024 * 1024  # 8MB
VOUCHER_RETENTION_DAYS = 365  # Current technical retention window for voucher PDFs.
ORDER_ACCESS_TOKEN_TTL_HOURS = 24
ORDER_ACCESS_TOKEN_HEADER = "X-Order-Access-Token"
ORDER_TRACK_LIMIT_WINDOW_SECONDS = 300
ORDER_TRACK_LIMIT_MAX_ATTEMPTS = 12
ORDER_TRACK_GLOBAL_LIMIT_WINDOW_SECONDS = 300
ORDER_TRACK_GLOBAL_LIMIT_MAX_ATTEMPTS = 40
ORDER_CONFIRM_LIMIT_WINDOW_SECONDS = 300
ORDER_CONFIRM_LIMIT_MAX_ATTEMPTS = 25

app = FastAPI(title="Tendinha do Saber API")
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger("tendinha")

# ---------- Helpers ----------
def now_utc() -> datetime:
    return datetime.now(timezone.utc)

def iso(dt: datetime) -> str:
    return dt.isoformat()

def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False

def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "role": role,
        "exp": now_utc() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MIN),
        "type": "access",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def strip_isbn(s: str) -> str:
    return re.sub(r"[^0-9Xx]", "", s or "")


# ---- Book identifier helpers (Bloco A) ----
# A book is identified by (in priority order): isbn13 (13 digits) OR slug OR
# pe_code (Porto Editora internal code, for books without ISBN). All three
# fields are unique when present (partial indexes in ensure_indexes()).
_SLUG_STRIP_RE = re.compile(r"[^a-z0-9]+")


def _slugify(text: str) -> str:
    """Deterministic ASCII slug from a title. Handles Portuguese accents,
    removes symbols, collapses whitespace to hyphens. Returns empty string
    if the input has no usable characters."""
    if not text:
        return ""
    import unicodedata
    # Strip diacritics: "Matemática" -> "Matematica"
    ascii_text = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode("ascii")
    ascii_text = ascii_text.lower().strip()
    # Replace anything non-alphanumeric with a hyphen, then collapse hyphens
    slug = _SLUG_STRIP_RE.sub("-", ascii_text).strip("-")
    return slug[:80]  # cap length to keep URLs sane


async def _ensure_unique_slug(base: str, exclude_id: Optional[str] = None) -> str:
    """Given a base slug, returns a slug guaranteed to be unique in the books
    collection. Adds -2, -3, ... suffix on collision. `exclude_id` lets a
    book keep its own slug during an update without triggering a suffix."""
    if not base:
        # Plan B: caller must supply a synthetic fallback (typically book UUID).
        return ""
    candidate = base
    n = 1
    while True:
        filt = {"slug": candidate}
        if exclude_id:
            filt["id"] = {"$ne": exclude_id}
        collision = await db.books.find_one(filt, {"_id": 0, "id": 1})
        if not collision:
            return candidate
        n += 1
        candidate = f"{base}-{n}"


def _book_key_filter(key: str) -> dict:
    """Builds a MongoDB filter that resolves a URL/API key to a single book.
    The key may be: (a) a 13-digit ISBN, (b) a slug, or (c) a PE code. We do
    NOT strip_isbn() the key blindly, because slugs contain hyphens/letters
    that would be destroyed. Instead we try each field verbatim, and also
    try the digits-only form for the isbn13 field so hyphenated ISBNs still
    resolve."""
    key = (key or "").strip()
    isbn_clean = strip_isbn(key)
    ors = [{"slug": key}, {"pe_code": key}]
    if isbn_clean:
        ors.append({"isbn13": isbn_clean})
    return {"$or": ors}


async def _find_book_by_key(key: str) -> Optional[dict]:
    return await db.books.find_one(_book_key_filter(key), {"_id": 0})


def gen_id() -> str:
    return str(uuid.uuid4())

def clean_doc(d: Optional[dict]) -> Optional[dict]:
    if d is None:
        return None
    d.pop("_id", None)
    return d


def _normalize_order_no(order_no: str) -> str:
    return (order_no or "").strip().upper()


def _new_order_access_token() -> str:
    return secrets.token_urlsafe(32)


def _hash_order_access_token(token: str) -> str:
    return hashlib.sha256((token or "").encode("utf-8")).hexdigest()


def _order_access_expires_at() -> datetime:
    return now_utc() + timedelta(hours=ORDER_ACCESS_TOKEN_TTL_HOURS)


def _request_origin_hint(request: Request) -> str:
    xff = (request.headers.get("X-Forwarded-For") or "").split(",")[0].strip()
    x_real_ip = (request.headers.get("X-Real-IP") or "").strip()
    remote = request.client.host if request.client else ""
    return xff or x_real_ip or remote or "anon"


def _hash_identifier(value: str) -> str:
    return hashlib.sha256((value or "").encode("utf-8")).hexdigest()


def _public_order_access_error() -> HTTPException:
    return HTTPException(404, "Encomenda não encontrada ou acesso inválido")


def _public_order_track_error() -> HTTPException:
    return HTTPException(404, "Não foi possível encontrar uma encomenda correspondente aos dados indicados.")


def _coerce_datetime(value: Any) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            return None
    return None


def _is_valid_order_access_token(order: dict, provided_token: Optional[str]) -> bool:
    if not provided_token:
        return False
    stored_hash = order.get("access_token_hash")
    expires_at = _coerce_datetime(order.get("access_token_expires_at"))
    if not stored_hash or not expires_at:
        return False
    if expires_at <= now_utc():
        return False
    provided_hash = _hash_order_access_token(provided_token)
    return hmac.compare_digest(str(stored_hash), provided_hash)


def _order_confirmation_payload(order: dict) -> dict:
    return {
        "order_no": order.get("order_no"),
        "status": order.get("status"),
        "created_at": order.get("created_at"),
        "items": [
            {
                "isbn13": it.get("isbn13"),
                "title": it.get("title"),
                "qty": it.get("qty"),
                "line_total": it.get("line_total"),
            }
            for it in (order.get("items") or [])
        ],
        "totals": {
            "subtotal_manuals": (order.get("totals") or {}).get("subtotal_manuals"),
            "subtotal_workbooks": (order.get("totals") or {}).get("subtotal_workbooks"),
            "discount_workbooks": (order.get("totals") or {}).get("discount_workbooks"),
            "lamination_total": (order.get("totals") or {}).get("lamination_total"),
            "bags_total": (order.get("totals") or {}).get("bags_total"),
            "shipping_cost": (order.get("totals") or {}).get("shipping_cost"),
            "total": (order.get("totals") or {}).get("total"),
        },
        "delivery": {
            "method": (order.get("delivery") or {}).get("method"),
            "concelho": (order.get("delivery") or {}).get("concelho"),
        },
        "customer": {
            "phone": (order.get("customer") or {}).get("phone"),
        },
    }


def _order_tracking_payload(order: dict) -> dict:
    return {
        "order_no": order.get("order_no"),
        "status": order.get("status"),
        "created_at": order.get("created_at"),
        "delivery": {
            "method": (order.get("delivery") or {}).get("method"),
            "concelho": (order.get("delivery") or {}).get("concelho"),
        },
        "totals": {
            "total": (order.get("totals") or {}).get("total"),
            "shipping_cost": (order.get("totals") or {}).get("shipping_cost"),
        },
        "items": [
            {
                "title": it.get("title"),
                "qty": it.get("qty"),
                "line_total": it.get("line_total"),
            }
            for it in (order.get("items") or [])
        ],
    }


async def _new_unique_order_no(max_attempts: int = 12) -> str:
    for _ in range(max_attempts):
        ts = int(now_utc().timestamp())
        suffix = secrets.token_hex(3).upper()
        candidate = f"TS-{ts}-{suffix}"
        exists = await db.orders.find_one({"order_no": candidate}, {"_id": 1})
        if not exists:
            return candidate
    raise HTTPException(500, "Não foi possível gerar número de encomenda único")


async def _check_order_access_rate_limit(scope: str, key_parts: List[str], max_attempts: int, window_seconds: int):
    now = now_utc()
    window_start = now - timedelta(seconds=window_seconds)
    raw = "|".join([scope] + [str(k or "") for k in key_parts])
    identifier_hash = _hash_identifier(raw.lower())
    count = await db.order_access_rate_limits.count_documents({
        "scope": scope,
        "identifier_hash": identifier_hash,
        "created_at_dt": {"$gte": window_start},
    })
    if count >= max_attempts:
        raise HTTPException(429, "Demasiadas tentativas. Tente novamente em instantes.")
    await db.order_access_rate_limits.insert_one({
        "scope": scope,
        "identifier_hash": identifier_hash,
        "created_at_dt": now,
        "expires_at": now + timedelta(seconds=window_seconds + 60),
    })

# ---------- Auth dependency ----------
async def get_current_user(request: Request) -> dict:
    auth = request.headers.get("Authorization", "")
    token = None
    if auth.startswith("Bearer "):
        token = auth[7:]
    if not token:
        token = request.cookies.get("access_token")
    if not token:
        raise HTTPException(401, "Não autenticado")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(401, "Token inválido")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(401, "Utilizador não encontrado")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Sessão expirada")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Token inválido")

async def get_current_user_optional(request: Request) -> Optional[dict]:
    """Same as get_current_user but never raises — returns None for
    anonymous requests. Used by public endpoints (voucher submission)
    that should still associate a customer_id when the caller happens
    to be logged in."""
    auth = request.headers.get("Authorization", "")
    token = None
    if auth.startswith("Bearer "):
        token = auth[7:]
    if not token:
        token = request.cookies.get("access_token")
    if not token:
        return None
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            return None
        return await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    except jwt.PyJWTError:
        return None

async def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") not in ("staff", "admin", "super_admin"):
        raise HTTPException(403, "Acesso restrito a administradores")
    return user

async def require_manager(user: dict = Depends(get_current_user)) -> dict:
    """Admin/Super Admin only — excludes 'staff'. Use for customer
    block/delete, financial reports, partners/promo-codes, content and
    settings management."""
    if user.get("role") not in ("admin", "super_admin"):
        raise HTTPException(403, "Acesso restrito a administradores com permissões de gestão")
    return user

async def require_super_admin(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != "super_admin":
        raise HTTPException(403, "Acesso restrito ao Super Administrador")
    return user

async def log_action(admin_id: str, action: str, entity: str, entity_id: Optional[str] = None, details: Optional[dict] = None):
    await db.activity_logs.insert_one({
        "id": gen_id(),
        "admin_id": admin_id,
        "action_type": action,
        "entity": entity,
        "entity_id": entity_id,
        "details": details or {},
        "timestamp": iso(now_utc()),
    })

# ---------- Pydantic models ----------
class RegisterIn(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str = Field(min_length=2)

class LoginIn(BaseModel):
    email: EmailStr
    password: str

class FirstLoginIn(BaseModel):
    token: str
    password: str = Field(min_length=8)

class ForgotIn(BaseModel):
    email: EmailStr

class ResetIn(BaseModel):
    token: str
    password: str = Field(min_length=6)

class ChangePasswordIn(BaseModel):
    current_password: str
    new_password: str = Field(min_length=8)

class BookIn(BaseModel):
    isbn13: Optional[str] = ""           # Bloco A: agora opcional (pode ser vazio se pe_code presente)
    pe_code: Optional[str] = None        # Bloco A: código interno Porto Editora (livros sem ISBN)
    slug: Optional[str] = None           # Bloco A: URL slug (gerado auto). SÓ visível internamente.
    related_book_id: Optional[str] = None  # Bloco A: futura ligação manual↔caderno (deixado vazio agora)
    title: str
    author: Optional[str] = ""
    publisher: Optional[str] = ""
    year: Optional[int] = None
    subject: Optional[str] = ""
    price: float
    type: str = "Manual"  # Manual | Workbook
    status: str = "Available"  # Available | PreOrder | Unavailable
    stock_qty: int = 0
    synopsis: Optional[str] = ""
    features: Optional[Dict[str, Any]] = {}
    image_url: Optional[str] = ""
    is_lamination_eligible: bool = True

class MunicipalityIn(BaseModel):
    name: str

class SchoolIn(BaseModel):
    name: str
    municipality_id: str
    grades_taught: Optional[List[str]] = None

class SchoolBookIn(BaseModel):
    school_id: str
    isbn13: str
    grade_level: str

class PartnerIn(BaseModel):
    name: str
    logo_url: Optional[str] = ""
    description: Optional[str] = ""
    promo_code: str
    discount_value: float = 5.0
    valid_from: Optional[str] = None
    valid_until: Optional[str] = None
    usage_limit: Optional[int] = None
    active: bool = True

class VoucherSubmitIn(BaseModel):
    code: Optional[str] = None
    pdf_url: Optional[str] = None
    notes: Optional[str] = None
    name: Optional[str] = None
    contact: Optional[str] = None
    manuals: Optional[str] = None
    wants_workbooks: Optional[bool] = False
    workbook_details: Optional[str] = None
    wants_lamination: Optional[bool] = False
    lamination_details: Optional[str] = None

class CartItem(BaseModel):
    isbn13: str
    qty: int = 1
    lamination: bool = False

class PromoValidateIn(BaseModel):
    items: List[CartItem]
    promo_code: Optional[str] = None
    bags_qty: Optional[int] = 0

class OrderCreateIn(BaseModel):
    items: List[CartItem]
    promo_code: Optional[str] = None
    customer_name: str
    customer_email: EmailStr
    customer_phone: str
    delivery_method: str  # hand_delivery | shipping
    delivery_concelho: Optional[str] = None  # Bloco B: concelho de Aveiro escolhido
    address: Optional[str] = ""
    postal_code: Optional[str] = ""
    notes: Optional[str] = ""
    # Bloco C: dados de faturação (opcionais; obrigatórios apenas quando o cliente pede fatura com NIF)
    wants_invoice: Optional[bool] = False
    nif: Optional[str] = None
    fiscal_name: Optional[str] = None
    bags_qty: Optional[int] = 0
    terms_accepted: Optional[bool] = False
    lamination_early_start_ack: Optional[bool] = False


class OrderTrackIn(BaseModel):
    order_no: str
    email: EmailStr

class SettingIn(BaseModel):
    lamination_price: Optional[float] = None
    aveiro_postcodes: Optional[List[str]] = None
    shipping_flat_rate: Optional[float] = None
    shipping_rates: Optional[Dict[str, float]] = None  # Bloco B: {concelho: preço}
    ctt_enabled: Optional[bool] = None                 # Bloco B: CTT preparado (default off)
    google_analytics_id: Optional[str] = None
    google_ads_id: Optional[str] = None
    facebook_pixel_id: Optional[str] = None
    google_site_verification: Optional[str] = None
    site_url: Optional[str] = None
    # Optional cover URL template for the publisher (e.g. Porto Editora / WOOK).
    # Use {isbn} as the placeholder, e.g. "https://.../{isbn}.jpg". Tried FIRST
    # when enriching covers, so when the publisher publishes a manual's cover it
    # is picked up automatically on the next "Procurar capas" run.
    publisher_cover_template: Optional[str] = None

class WishlistIn(BaseModel):
    isbn13: str

class CategoryIn(BaseModel):
    name: str
    is_active: bool = True

class ContentIn(BaseModel):
    about_us: Optional[str] = None
    hero_title: Optional[str] = None
    hero_subtitle: Optional[str] = None
    footer_text: Optional[str] = None
    instagram_handle: Optional[str] = None
    instagram_url: Optional[str] = None
    partners_cta: Optional[str] = None
    promotions_label: Optional[str] = None

# ---------- Auth routes ----------
@api.post("/auth/register")
async def register(payload: RegisterIn):
    email = payload.email.lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(400, "Já existe uma conta com este email")
    user = {
        "id": gen_id(),
        "email": email,
        "name": payload.name,
        "role": "customer",
        "password_hash": hash_password(payload.password),
        "created_at": iso(now_utc()),
    }
    await db.users.insert_one(user)
    token = create_access_token(user["id"], email, "customer")
    return {"token": token, "user": {k: v for k, v in user.items() if k not in ("password_hash", "_id")}}

@api.post("/auth/login")
async def login(payload: LoginIn, request: Request):
    email = payload.email.lower()
    identifier = email  # K8s ingress rotates client IPs; use email alone for reliable lockout

    # Brute force check
    rec = await db.login_attempts.find_one({"identifier": identifier})
    if rec and rec.get("locked_until"):
        locked = datetime.fromisoformat(rec["locked_until"])
        if locked > now_utc():
            mins = int((locked - now_utc()).total_seconds() / 60) + 1
            raise HTTPException(429, f"Conta bloqueada. Tente novamente em {mins} minutos.")

    user = await db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user["password_hash"]):
        # increment attempts
        attempts = (rec.get("attempts", 0) if rec else 0) + 1
        update = {"identifier": identifier, "attempts": attempts, "last_at": iso(now_utc())}
        if attempts >= 5:
            update["locked_until"] = iso(now_utc() + timedelta(minutes=30))
            update["attempts"] = 0
        await db.login_attempts.update_one({"identifier": identifier}, {"$set": update}, upsert=True)
        raise HTTPException(401, "Credenciais inválidas")

    if user.get("is_blocked"):
        raise HTTPException(403, "Esta conta foi bloqueada. Contacte a Tendinha do Saber para mais informações.")

    # If admin has pending first-login flow
    if user.get("must_change_password"):
        raise HTTPException(403, "Primeiro acesso: defina uma nova palavra-passe através do link recebido por email.")

    await db.login_attempts.delete_one({"identifier": identifier})
    token = create_access_token(user["id"], user["email"], user["role"])
    safe = {k: v for k, v in user.items() if k not in ("_id", "password_hash")}
    return {"token": token, "user": safe}

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user

@api.post("/auth/forgot-password")
async def forgot(payload: ForgotIn):
    user = await db.users.find_one({"email": payload.email.lower()})
    # Always success to prevent enumeration
    if user:
        token = secrets.token_urlsafe(32)
        await db.password_reset_tokens.insert_one({
            "id": gen_id(),
            "token": token,
            "user_id": user["id"],
            "used": False,
            "expires_at": iso(now_utc() + timedelta(hours=1)),
        })
        # MOCKED EMAIL: log link
        logger.info(f"[MOCKED EMAIL] Password reset for {payload.email}: token={token}")
    return {"message": "Se o email existir, receberá instruções de recuperação."}

@api.post("/auth/reset-password")
async def reset(payload: ResetIn):
    rec = await db.password_reset_tokens.find_one({"token": payload.token, "used": False})
    if not rec or datetime.fromisoformat(rec["expires_at"]) < now_utc():
        raise HTTPException(400, "Token inválido ou expirado")
    await db.users.update_one(
        {"id": rec["user_id"]},
        {"$set": {"password_hash": hash_password(payload.password), "must_change_password": False}},
    )
    await db.password_reset_tokens.update_one({"id": rec["id"]}, {"$set": {"used": True}})
    return {"message": "Palavra-passe atualizada com sucesso"}

@api.post("/auth/first-login")
async def first_login(payload: FirstLoginIn):
    rec = await db.password_reset_tokens.find_one({"token": payload.token, "used": False})
    if not rec or datetime.fromisoformat(rec["expires_at"]) < now_utc():
        raise HTTPException(400, "Token inválido ou expirado")
    await db.users.update_one(
        {"id": rec["user_id"]},
        {"$set": {"password_hash": hash_password(payload.password), "must_change_password": False}},
    )
    await db.password_reset_tokens.update_one({"id": rec["id"]}, {"$set": {"used": True}})
    user = await db.users.find_one({"id": rec["user_id"]}, {"_id": 0, "password_hash": 0})
    token = create_access_token(user["id"], user["email"], user["role"])
    return {"token": token, "user": user}

@api.post("/auth/change-password")
async def change_password(payload: ChangePasswordIn, user: dict = Depends(get_current_user)):
    stored = await db.users.find_one({"id": user["id"]})
    if not stored or not verify_password(payload.current_password, stored.get("password_hash", "")):
        raise HTTPException(400, "A password atual está incorreta")
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"password_hash": hash_password(payload.new_password)}},
    )
    return {"message": "Password atualizada com sucesso"}

# ---------- Books ----------
@api.get("/books")
async def list_books(
    q: Optional[str] = None,
    subject: Optional[str] = None,
    type: Optional[str] = None,
    status: Optional[str] = None,
    school_id: Optional[str] = None,
    concelho: Optional[str] = None,
    school_name: Optional[str] = None,
    grade_level: Optional[str] = None,
    stock: Optional[str] = None,
    limit: int = 20,
    skip: int = 0,
    page: Optional[int] = None,
):
    filt: Dict[str, Any] = {}
    if q:
        q_clean = strip_isbn(q)
        regex = {"$regex": re.escape(q), "$options": "i"}
        ors = [{"title": regex}, {"author": regex}, {"subject": regex}, {"publisher": regex}, {"pe_code": regex}]
        if q_clean:
            ors.append({"isbn13": q_clean})
        filt["$or"] = ors
    if subject:
        filt["subject"] = subject
    if type:
        filt["type"] = type
    if status:
        filt["status"] = status

    if stock == "low":
        filt["stock_qty"] = {"$lte": STOCK_LOW_THRESHOLD}
    elif stock == "high":
        filt["stock_qty"] = {"$gt": STOCK_LOW_THRESHOLD}

    adoption_isbns: List[str] = []
    if concelho and school_name:
        active_year = await _get_active_school_year()
        if active_year:
            adoption_filter: Dict[str, Any] = {
                "school_year": active_year,
                "concelho": concelho,
                "escola": school_name,
            }
            if grade_level:
                adoption_filter["grade"] = grade_level
            adoption_isbns = await db.school_adoptions.distinct("isbn13", adoption_filter)
            filt["isbn13"] = {"$in": adoption_isbns}
    elif school_id:
        school = await db.schools.find_one({"id": school_id}, {"_id": 0, "name": 1, "municipality_id": 1})
        if school:
            mun = await db.municipalities.find_one({"id": school.get("municipality_id")}, {"_id": 0, "name": 1})
            active_year = await _get_active_school_year()
            if mun and active_year:
                adoption_filter = {
                    "school_year": active_year,
                    "concelho": mun.get("name"),
                    "escola": school.get("name"),
                }
                if grade_level:
                    adoption_filter["grade"] = grade_level
                adoption_isbns = await db.school_adoptions.distinct("isbn13", adoption_filter)
        if adoption_isbns:
            filt["isbn13"] = {"$in": adoption_isbns}
        else:
            sb_filter: Dict[str, Any] = {"school_id": school_id}
            if grade_level:
                sb_filter["grade_level"] = grade_level
            isbns = await db.school_books.distinct("isbn13", sb_filter)
            filt["isbn13"] = {"$in": isbns}

    limit = max(1, min(limit, 500))
    if page is not None:
        skip = max(0, (page - 1)) * limit

    total = await db.books.count_documents(filt)
    cursor = db.books.find(filt, {"_id": 0}).skip(skip).limit(limit)
    items = await cursor.to_list(length=limit)
    return {
        "items": items,
        "total": total,
        "page": (skip // limit) + 1 if limit else 1,
        "page_size": limit,
        "pages": max(1, (total + limit - 1) // limit),
    }

@api.get("/books/subjects")
async def list_subjects():
    return await db.books.distinct("subject")

@api.get("/books/{key}")
async def get_book(key: str):
    """Bloco A: `key` pode ser ISBN-13, slug ou código PE."""
    book = await _find_book_by_key(key)
    if not book:
        raise HTTPException(404, "Livro não encontrado")
    return book

@api.post("/admin/books")
async def create_book(payload: BookIn, admin: dict = Depends(require_admin)):
    isbn = strip_isbn(payload.isbn13 or "")
    pe_code = (payload.pe_code or "").strip() or None
    if not isbn and not pe_code:
        raise HTTPException(400, "Indique um ISBN OU um Código PE (pelo menos um é obrigatório).")
    if isbn and len(isbn) != 13:
        raise HTTPException(400, "ISBN inválido — deve ter 13 dígitos.")
    if isbn and await db.books.find_one({"isbn13": isbn}):
        raise HTTPException(400, "ISBN já existe")
    if pe_code and await db.books.find_one({"pe_code": pe_code}):
        raise HTTPException(400, "Código PE já existe")

    doc = payload.model_dump()
    doc["isbn13"] = isbn or ""
    doc["pe_code"] = pe_code
    doc["id"] = gen_id()
    # Slug: se admin forneceu um explícito, respeita-o; senão gera do título.
    # Plan B: se título vazio ou slug vazio, usa o UUID como slug para não partir.
    provided_slug = _slugify(payload.slug or "")
    base = provided_slug or _slugify(payload.title) or doc["id"]
    doc["slug"] = await _ensure_unique_slug(base, exclude_id=doc["id"])
    doc["created_at"] = iso(now_utc())
    await db.books.insert_one(doc)
    await log_action(admin["id"], "create", "book", doc["id"], {"isbn": doc["isbn13"], "pe_code": pe_code, "slug": doc["slug"]})
    doc.pop("_id", None)
    return doc

@api.put("/admin/books/{key}")
async def update_book(key: str, payload: BookIn, admin: dict = Depends(require_admin)):
    """Bloco A: `key` pode ser ISBN-13, slug ou código PE."""
    existing = await _find_book_by_key(key)
    if not existing:
        raise HTTPException(404, "Livro não encontrado")

    update = payload.model_dump()
    new_isbn = strip_isbn(update.get("isbn13") or "")
    new_pe = (update.get("pe_code") or "").strip() or None
    if not new_isbn and not new_pe:
        raise HTTPException(400, "Indique um ISBN OU um Código PE (pelo menos um é obrigatório).")
    if new_isbn and len(new_isbn) != 13:
        raise HTTPException(400, "ISBN inválido — deve ter 13 dígitos.")
    # Impedir colisões noutros livros (mas permitir manter o próprio)
    if new_isbn:
        clash = await db.books.find_one({"isbn13": new_isbn, "id": {"$ne": existing["id"]}}, {"_id": 0, "id": 1})
        if clash:
            raise HTTPException(400, "ISBN já pertence a outro livro")
    if new_pe:
        clash = await db.books.find_one({"pe_code": new_pe, "id": {"$ne": existing["id"]}}, {"_id": 0, "id": 1})
        if clash:
            raise HTTPException(400, "Código PE já pertence a outro livro")

    update["isbn13"] = new_isbn or ""
    update["pe_code"] = new_pe

    # Slug: se admin escreveu um slug diferente, respeita; senão mantém o atual;
    # se atualmente vazio, gera-o agora a partir do título.
    incoming_slug = _slugify(update.get("slug") or "")
    current_slug = existing.get("slug") or ""
    if incoming_slug and incoming_slug != current_slug:
        update["slug"] = await _ensure_unique_slug(incoming_slug, exclude_id=existing["id"])
    elif not current_slug:
        base = _slugify(update.get("title") or existing.get("title") or "") or existing["id"]
        update["slug"] = await _ensure_unique_slug(base, exclude_id=existing["id"])
    else:
        update["slug"] = current_slug

    await db.books.update_one({"id": existing["id"]}, {"$set": update})
    await log_action(admin["id"], "update", "book", existing["id"])
    return {"ok": True}

@api.delete("/admin/books/{key}")
async def delete_book(key: str, admin: dict = Depends(require_admin)):
    """Bloco A: `key` pode ser ISBN-13, slug ou código PE."""
    existing = await _find_book_by_key(key)
    if not existing:
        raise HTTPException(404, "Livro não encontrado")
    res = await db.books.delete_one({"id": existing["id"]})
    if res.deleted_count == 0:
        raise HTTPException(404, "Livro não encontrado")
    await log_action(admin["id"], "delete", "book", existing["id"], {"isbn": existing.get("isbn13"), "pe_code": existing.get("pe_code")})
    return {"ok": True}

# ---------- Excel Import ----------
def _normalize_import_df(content: bytes):
    """Reads the Excel and normalizes column names to the internal schema.
    Returns the dataframe. Raises HTTPException(400) on unreadable files."""
    try:
        df = pd.read_excel(BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Falha ao ler ficheiro Excel: {e}")
    df.columns = [str(c).strip().lower() for c in df.columns]
    col_map = {
        "ciclo": "cycle", "ciclo de ensino": "cycle",
        "ano": "grade_level", "ano de escolaridade": "grade_level",
        "disciplina": "subject", "editora": "publisher",
        "título": "title", "titulo": "title",
        "isbn": "isbn13", "artigo": "type", "pvp": "price",
        "preço": "price", "preco": "price",
        "autor": "author", "autor(es)": "author", "autores": "author",
        # Bloco A: Código Porto Editora (para livros sem ISBN)
        "código pe": "pe_code", "codigo pe": "pe_code",
        "código porto editora": "pe_code", "codigo porto editora": "pe_code",
        "codigo_pe": "pe_code", "código_pe": "pe_code",
        "cod pe": "pe_code", "cod. pe": "pe_code",
    }
    df.rename(columns={k: v for k, v in col_map.items() if k in df.columns}, inplace=True)
    return df


def _classify_import_rows(df) -> List[dict]:
    """Validates and classifies every row WITHOUT touching the database.
    Each returned record has an 'action' of 'new' | 'update' | 'error' and,
    for valid rows, a normalized 'data' dict ready to be persisted.

    Bloco A: aceita linhas SEM ISBN se tiverem Código PE + Título + Preço.
    Rejeita linhas sem NENHUM identificador (nem ISBN nem Código PE)."""

    def _clean_cell(v) -> str:
        """Pandas devolve NaN/None em células vazias e converte para 'nan' quando
        forçado a str. Este helper trata NaN/None como vazio, e remove um '.0'
        residual de códigos numéricos que o pandas leu como float (perdendo,
        infelizmente, zeros à esquerda — cabe ao utilizador formatar a coluna
        como Texto no Excel se quiser preservá-los)."""
        if v is None:
            return ""
        try:
            # pd.isna funciona para NaN, NaT e None
            import pandas as _pd
            if _pd.isna(v):
                return ""
        except Exception:
            pass
        s = str(v).strip()
        if s.lower() in ("nan", "none", "nat"):
            return ""
        if s.endswith(".0") and s[:-2].isdigit():
            s = s[:-2]
        return s

    rows: List[dict] = []
    for idx, row in df.iterrows():
        line_no = int(idx) + 2  # +2: header row + 1-based for humans
        isbn = strip_isbn(_clean_cell(row.get("isbn13", "")))
        pe_code = _clean_cell(row.get("pe_code", ""))
        title = _clean_cell(row.get("title", ""))
        try:
            price_raw = row.get("price", 0)
            price = float(price_raw) if price_raw not in (None, "") else 0.0
        except Exception:
            price = 0.0

        problems = []
        # Bloco A: ISBN OU Código PE (não os dois obrigatórios)
        has_isbn = len(isbn) == 13
        if not has_isbn and not pe_code:
            problems.append("é preciso ISBN (13 dígitos) OU Código PE")
        elif isbn and not has_isbn:
            problems.append("ISBN inválido (deve ter 13 dígitos)")
        if price <= 0:
            problems.append("preço em falta ou ≤ 0")
        if not title:
            problems.append("título em falta")

        if problems:
            rows.append({
                "line": line_no,
                "isbn": isbn,
                "pe_code": pe_code,
                "title": title,
                "action": "error",
                "issue": "; ".join(problems),
            })
            continue

        # Bloco A: detecta cadernos por coluna "artigo/tipo" OU por título
        artigo_field = _clean_cell(row.get("type", "")).lower()
        title_lower = title.lower()
        is_workbook = (
            "caderno" in artigo_field or "workbook" in artigo_field
            or "fichas" in artigo_field or "caderno" in title_lower
        )
        item_type = "Workbook" if is_workbook else "Manual"

        rows.append({
            "line": line_no,
            "isbn": isbn,
            "pe_code": pe_code,
            "title": title,
            "action": "pending",  # resolved to new/update during preview against DB
            "data": {
                "isbn13": isbn if has_isbn else "",
                "pe_code": pe_code or None,
                "title": title,
                "author": _clean_cell(row.get("author", "")),
                "publisher": _clean_cell(row.get("publisher", "")),
                "subject": _clean_cell(row.get("subject", "")),
                "price": round(price, 2),
                "type": item_type,
                "cycle": _clean_cell(row.get("cycle", "")),
                "grade_level": _clean_cell(row.get("grade_level", "")),
            },
        })
    return rows


@api.post("/admin/books/import/preview")
async def import_books_preview(file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    """PHASE 1 — dry run. Parses and validates the Excel, reports how many
    rows would be created / updated / rejected, and returns the validated
    payload (base64) to be sent back to /commit. The database is NOT touched."""
    content = await file.read()
    df = _normalize_import_df(content)
    classified = _classify_import_rows(df)

    valid_rows = []
    new_count = update_count = 0
    for r in classified:
        if r["action"] == "error":
            continue
        # Bloco A: procura por ISBN OU por Código PE
        d = r["data"]
        existing = None
        if d.get("isbn13"):
            existing = await db.books.find_one({"isbn13": d["isbn13"]}, {"_id": 0, "id": 1})
        if not existing and d.get("pe_code"):
            existing = await db.books.find_one({"pe_code": d["pe_code"]}, {"_id": 0, "id": 1})
        r["action"] = "update" if existing else "new"
        if existing:
            update_count += 1
        else:
            new_count += 1
        valid_rows.append(r)

    errors = [r for r in classified if r["action"] == "error"]
    # Token carries only the validated rows, so /commit re-validates nothing
    # the user didn't already see in the preview.
    import base64 as _b64
    token = _b64.b64encode(json.dumps([r["data"] for r in valid_rows]).encode()).decode()

    return {
        "summary": {
            "new": new_count,
            "update": update_count,
            "errors": len(errors),
            "total": len(classified),
        },
        "preview": (
            [{"line": r["line"], "isbn": r["isbn"], "title": r["title"], "action": r["action"]} for r in valid_rows[:100]]
            + [{"line": r["line"], "isbn": r["isbn"], "title": r["title"], "action": "error", "issue": r["issue"]} for r in errors[:100]]
        ),
        "errors": [{"line": r["line"], "isbn": r["isbn"], "title": r["title"], "issue": r["issue"]} for r in errors],
        "commit_token": token,
    }


class ImportCommitIn(BaseModel):
    commit_token: str


@api.post("/admin/books/import/commit")
async def import_books_commit(payload: ImportCommitIn, admin: dict = Depends(require_admin)):
    """PHASE 2 — apply. Takes the validated payload from /preview and performs
    the UPSERT. Existing ISBNs update price/type/title/publisher/author/subject
    but preserve manually-curated synopsis and cover image. Invalid rows never
    reach this stage."""
    import base64 as _b64
    try:
        rows = json.loads(_b64.b64decode(payload.commit_token).decode())
    except Exception:
        raise HTTPException(400, "Token de importação inválido. Repita a pré-visualização.")
    if not isinstance(rows, list):
        raise HTTPException(400, "Token de importação corrompido.")

    created = updated = 0
    for data in rows:
        isbn = strip_isbn(str(data.get("isbn13", "") or ""))
        pe_code = (data.get("pe_code") or "").strip() or None
        try:
            price = float(data.get("price", 0) or 0)
        except Exception:
            price = 0.0
        # Bloco A: aceita rows com ISBN válido OU com Código PE
        has_isbn = len(isbn) == 13
        if (not has_isbn and not pe_code) or price <= 0:
            continue  # defensive: invalid rows never persist

        # Localizar existente por ISBN OU por Código PE
        existing = None
        if has_isbn:
            existing = await db.books.find_one({"isbn13": isbn})
        if not existing and pe_code:
            existing = await db.books.find_one({"pe_code": pe_code})

        if existing:
            update_set = {
                "price": round(price, 2),
                "type": data.get("type", "Manual"),
                "title": data.get("title") or existing.get("title", ""),
                "publisher": data.get("publisher", existing.get("publisher", "")),
                "author": data.get("author", existing.get("author", "")),
                "subject": data.get("subject", existing.get("subject", "")),
                "updated_at": iso(now_utc()),
            }
            # Se a linha traz ISBN e o existente não tinha, adiciona
            if has_isbn and not existing.get("isbn13"):
                update_set["isbn13"] = isbn
            # Se a linha traz pe_code e o existente não tinha, adiciona
            if pe_code and not existing.get("pe_code"):
                update_set["pe_code"] = pe_code
            await db.books.update_one({"id": existing["id"]}, {"$set": update_set})
            updated += 1
        else:
            new_id = gen_id()
            base_slug = _slugify(data.get("title") or "") or new_id
            slug = await _ensure_unique_slug(base_slug, exclude_id=new_id)
            doc = {
                "id": new_id,
                "isbn13": isbn if has_isbn else "",
                "pe_code": pe_code,
                "slug": slug,
                "related_book_id": None,
                "title": data.get("title", ""),
                "author": data.get("author", ""),
                "publisher": data.get("publisher", ""),
                "subject": data.get("subject", ""),
                "price": round(price, 2),
                "type": data.get("type", "Manual"),
                "status": "Available",
                "stock_qty": 0,
                "synopsis": "",
                "features": {"cycle": data.get("cycle", ""), "grade": data.get("grade_level", "")},
                "image_url": "",
                "is_lamination_eligible": True,
                "created_at": iso(now_utc()),
            }
            await db.books.insert_one(doc)
            created += 1

    await log_action(admin["id"], "import", "books", None, {"created": created, "updated": updated})
    return {"created": created, "updated": updated}


# ============================================================
# BLOCO B — Exportar livros para Excel
# ============================================================
@api.get("/admin/books/export")
async def export_books_xlsx(admin: dict = Depends(require_admin)):
    """Exporta TODOS os livros para .xlsx (openpyxl). ISBN e Código PE ficam
    formatados como texto para preservar zeros à esquerda (05000072 fica intacto)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Livros"
    headers = [
        "ISBN", "Código PE", "Slug", "Título", "Autor(es)", "Editora",
        "Disciplina", "Ano", "Ciclo de Ensino", "Tipo/Artigo",
        "PVP (€)", "Stock", "URL da capa", "ID interno",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    # Forçar coluna ISBN e Código PE como TEXTO para preservar zeros à esquerda
    ws.column_dimensions["A"].number_format = "@"  # ISBN
    ws.column_dimensions["B"].number_format = "@"  # Código PE

    async for b in db.books.find({}, {"_id": 0}).sort("title", 1):
        features = b.get("features") or {}
        row = [
            str(b.get("isbn13") or ""),
            str(b.get("pe_code") or ""),
            str(b.get("slug") or ""),
            b.get("title") or "",
            b.get("author") or "",
            b.get("publisher") or "",
            b.get("subject") or "",
            b.get("year") or features.get("grade") or "",
            features.get("cycle") or "",
            "Caderno" if b.get("type") == "Workbook" else "Manual",
            float(b.get("price") or 0),
            int(b.get("stock_qty") or 0),
            b.get("image_url") or "",
            b.get("id") or "",
        ]
        ws.append(row)
        # Reforçar formato texto célula a célula (algumas versões do Excel ignoram
        # o formato de coluna se o valor "parece" numérico)
        r = ws.max_row
        ws.cell(row=r, column=1).number_format = "@"
        ws.cell(row=r, column=2).number_format = "@"

    # Auto-fit razoável nas colunas
    for i, col in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[chr(64 + i)].width = min(max(max_len + 2, 12), 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    await log_action(admin["id"], "export", "books", None, {"count": ws.max_row - 1})
    filename = f"livros-tendinha-{now_utc().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ============================================================
# BLOCO C — Importação de ligações manual↔caderno
# ============================================================
@api.post("/admin/books/links/import")
async def import_book_links(file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    """Importa ligações caderno→manual a partir de um Excel com uma folha "Ligacoes".
    Colunas: Código PE (caderno) | Caderno | Ano | Disciplina | ISBN (manual) | Manual | Confianca
    Repetível — corrida nova atualiza (não duplica).
    """
    content = await file.read()
    try:
        # Tentar folha "Ligacoes"; se não existir, usa a primeira
        try:
            df = pd.read_excel(BytesIO(content), sheet_name="Ligacoes")
        except Exception:
            df = pd.read_excel(BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Falha ao ler ficheiro: {e}")

    df.columns = [str(c).strip().lower() for c in df.columns]
    col_map = {
        "código pe (caderno)": "pe_code", "codigo pe (caderno)": "pe_code",
        "código pe": "pe_code", "codigo pe": "pe_code",
        "isbn (manual)": "isbn13", "isbn": "isbn13",
    }
    df.rename(columns={k: v for k, v in col_map.items() if k in df.columns}, inplace=True)

    linked = 0
    skipped: List[dict] = []
    for idx, row in df.iterrows():
        line_no = int(idx) + 2

        def clean(v):
            if v is None: return ""
            try:
                import pandas as _pd
                if _pd.isna(v): return ""
            except Exception:
                pass
            s = str(v).strip()
            if s.lower() in ("nan", "none", "nat"): return ""
            if s.endswith(".0") and s[:-2].isdigit(): s = s[:-2]
            return s

        pe = clean(row.get("pe_code", ""))
        isbn = strip_isbn(clean(row.get("isbn13", "")))
        if not pe or not isbn:
            skipped.append({"line": line_no, "reason": "sem pe_code ou sem isbn"})
            continue
        caderno = await db.books.find_one({"pe_code": pe}, {"_id": 0, "id": 1, "title": 1})
        manual = await db.books.find_one({"isbn13": isbn}, {"_id": 0, "id": 1, "title": 1})
        if not caderno:
            skipped.append({"line": line_no, "reason": f"caderno com pe_code '{pe}' não existe"})
            continue
        if not manual:
            skipped.append({"line": line_no, "reason": f"manual com ISBN '{isbn}' não existe"})
            continue
        # Ligação bidirecional. O caderno guarda o ID do manual e vice-versa.
        # Antes de escrever, limpar ligações reversas obsoletas noutros livros
        # que apontassem para este par (caso a associação tenha mudado numa
        # corrida anterior). Torna o import verdadeiramente idempotente.
        await db.books.update_many(
            {"related_book_id": manual["id"], "id": {"$nin": [caderno["id"], manual["id"]]}},
            {"$unset": {"related_book_id": ""}},
        )
        await db.books.update_many(
            {"related_book_id": caderno["id"], "id": {"$nin": [caderno["id"], manual["id"]]}},
            {"$unset": {"related_book_id": ""}},
        )
        await db.books.update_one({"id": caderno["id"]}, {"$set": {"related_book_id": manual["id"]}})
        await db.books.update_one({"id": manual["id"]}, {"$set": {"related_book_id": caderno["id"]}})
        linked += 1

    await log_action(admin["id"], "import", "book_links", None, {"linked": linked, "skipped": len(skipped)})
    return {"linked": linked, "skipped_count": len(skipped), "skipped": skipped[:200]}


# ============================================================
# BLOCO C — Sugestão de caderno no carrinho
# ============================================================
@api.post("/cart/related-workbooks")
async def cart_related_workbooks(payload: PromoValidateIn):
    """Dado um carrinho (items com isbn13/slug/pe_code), devolve os cadernos
    associados aos manuais que estão no carrinho — SEM incluir cadernos que
    o cliente já tem no carrinho. Só sugestão, nunca automático."""
    # 1. Resolver cada item para o livro real, e coletar ids em carrinho
    in_cart_ids: set = set()
    manuals_in_cart: List[dict] = []
    for it in payload.items:
        book = await _find_book_by_key(it.isbn13)
        if not book:
            continue
        in_cart_ids.add(book["id"])
        if book.get("type") == "Manual" and book.get("related_book_id"):
            manuals_in_cart.append(book)
    # 2. Para cada manual, buscar caderno associado
    suggestions: List[dict] = []
    seen: set = set()
    for m in manuals_in_cart:
        related_id = m.get("related_book_id")
        if not related_id or related_id in in_cart_ids or related_id in seen:
            continue
        caderno = await db.books.find_one(
            {"id": related_id, "type": "Workbook", "status": {"$ne": "Unavailable"}},
            {"_id": 0},
        )
        if not caderno:
            continue
        seen.add(related_id)
        suggestions.append({
            "manual_id": m["id"],
            "manual_title": m.get("title"),
            "workbook": {
                "isbn13": caderno.get("isbn13") or "",
                "slug": caderno.get("slug") or "",
                "id": caderno["id"],
                "title": caderno.get("title", ""),
                "author": caderno.get("author", ""),
                "publisher": caderno.get("publisher", ""),
                "price": float(caderno.get("price", 0) or 0),
                "image_url": caderno.get("image_url", ""),
            },
        })
    return {"suggestions": suggestions}


# ============================================================
# BLOCO D — Adoções escolares por ano letivo
# ============================================================
class AdoptionsSchoolYearIn(BaseModel):
    school_year: Optional[str] = None


async def _get_active_school_year() -> Optional[str]:
    """Ano letivo actualmente activo (o que o site público mostra)."""
    settings = await db.settings.find_one({"id": "global"}, {"_id": 0, "adoptions_active_year": 1}) or {}
    year = settings.get("adoptions_active_year")
    if year:
        return year
    # Fallback: pega o ano letivo mais recente disponível em adoções
    doc = await db.school_adoptions.find_one({}, sort=[("school_year", -1)])
    return doc.get("school_year") if doc else None


@api.get("/admin/adoptions/years")
async def admin_list_adoption_years(admin: dict = Depends(require_admin)):
    years = await db.school_adoptions.distinct("school_year")
    years.sort(reverse=True)
    counts = {}
    for y in years:
        counts[y] = await db.school_adoptions.count_documents({"school_year": y})
    active = await _get_active_school_year()
    return {"years": [{"year": y, "count": counts[y]} for y in years], "active": active}


@api.put("/admin/adoptions/active-year")
async def admin_set_active_year(payload: AdoptionsSchoolYearIn, admin: dict = Depends(require_manager)):
    """Escolhe qual o ano letivo activo (o que o site público mostra)."""
    year = (payload.school_year or "").strip() or None
    await db.settings.update_one(
        {"id": "global"},
        {"$set": {"adoptions_active_year": year, "id": "global"}},
        upsert=True,
    )
    await log_action(admin["id"], "update", "adoptions_active_year", None, {"year": year})
    return {"active": year}


@api.post("/admin/adoptions/import")
async def import_adoptions(
    file: UploadFile = File(...),
    school_year: str = Form(...),
    admin: dict = Depends(require_admin),
):
    """Importa adoções DGE. Corrida com o mesmo `school_year` SUBSTITUI só esse
    ano letivo (delete_many restrito ao ano) — não toca noutros anos guardados."""
    if not school_year or "/" not in school_year:
        raise HTTPException(400, "Indique o ano letivo no formato AAAA/AAAA, ex: 2026/2027")
    content = await file.read()
    try:
        df = pd.read_excel(BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Falha ao ler Excel: {e}")

    df.columns = [str(c).strip().lower() for c in df.columns]
    col_map = {
        "código escola": "codigo_escola", "codigo escola": "codigo_escola",
        "ano": "grade", "ano de escolaridade": "grade",
        "título": "title", "titulo": "title",
    }
    df.rename(columns={k: v for k, v in col_map.items() if k in df.columns}, inplace=True)

    def clean(v):
        if v is None: return ""
        try:
            import pandas as _pd
            if _pd.isna(v): return ""
        except Exception: pass
        s = str(v).strip()
        if s.lower() in ("nan", "none", "nat"): return ""
        if s.endswith(".0") and s[:-2].isdigit(): s = s[:-2]
        return s

    # 1) Apagar SÓ este ano letivo (nunca toca noutros)
    prev = await db.school_adoptions.count_documents({"school_year": school_year})
    if prev > 0:
        await db.school_adoptions.delete_many({"school_year": school_year})

    # 2) Reunir catálogo pelo ISBN uma vez para performance
    known_isbns = set(await db.books.distinct("isbn13", {"isbn13": {"$type": "string", "$gt": ""}}))

    # 3) Inserir por lotes (mantém memória controlada e não bloqueia loop)
    BATCH = 1000
    docs: List[dict] = []
    total = 0
    matched = 0
    missing: dict = {}  # isbn -> title (para reporte)

    for _, row in df.iterrows():
        isbn = strip_isbn(clean(row.get("isbn13", "") or row.get("isbn", "")))
        if len(isbn) != 13:
            continue
        title = clean(row.get("title", ""))
        in_catalog = isbn in known_isbns
        if in_catalog:
            matched += 1
        else:
            if isbn not in missing:
                missing[isbn] = title
        docs.append({
            "school_year": school_year,
            "concelho": clean(row.get("concelho", "")),
            "agrupamento": clean(row.get("agrupamento", "")),
            "codigo_escola": clean(row.get("codigo_escola", "")),
            "escola": clean(row.get("escola", "")),
            "grade": clean(row.get("grade", "")),
            "subject": clean(row.get("disciplina", "")),
            "isbn13": isbn,
            "title": title,
            "publisher": clean(row.get("editora", "")),
            "in_catalog": in_catalog,
        })
        total += 1
        if len(docs) >= BATCH:
            await db.school_adoptions.insert_many(docs)
            docs = []
    if docs:
        await db.school_adoptions.insert_many(docs)

    # 4) Se este é o único ano com dados, activa-o
    settings = await db.settings.find_one({"id": "global"}, {"_id": 0, "adoptions_active_year": 1}) or {}
    if not settings.get("adoptions_active_year"):
        await db.settings.update_one(
            {"id": "global"},
            {"$set": {"adoptions_active_year": school_year, "id": "global"}},
            upsert=True,
        )

    await log_action(admin["id"], "import", "adoptions", None, {"school_year": school_year, "total": total, "matched": matched, "missing": len(missing)})
    return {
        "school_year": school_year,
        "previous_replaced": prev,
        "total": total,
        "matched": matched,
        "missing_count": len(missing),
        "missing_sample": [{"isbn13": k, "title": v} for k, v in list(missing.items())[:100]],
    }


# ---------- Adoções — endpoints públicos (para o dropdown em cascata) ----------
@api.get("/adoptions/concelhos")
async def adoptions_concelhos():
    year = await _get_active_school_year()
    if not year:
        return {"concelhos": [], "active_year": None}
    concelhos = await db.school_adoptions.distinct("concelho", {"school_year": year})
    concelhos = sorted([c for c in concelhos if c])
    return {"concelhos": concelhos, "active_year": year}


@api.get("/adoptions/schools")
async def adoptions_schools(concelho: str):
    year = await _get_active_school_year()
    if not year:
        return {"schools": [], "active_year": None}
    schools = await db.school_adoptions.distinct(
        "escola", {"school_year": year, "concelho": concelho}
    )
    schools = sorted([s for s in schools if s])
    return {"schools": schools, "active_year": year}


@api.get("/adoptions/grades")
async def adoptions_grades(concelho: str, escola: str):
    year = await _get_active_school_year()
    if not year:
        return {"grades": [], "active_year": None}
    grades = await db.school_adoptions.distinct(
        "grade", {"school_year": year, "concelho": concelho, "escola": escola}
    )
    # Ordenação natural (1.º, 2.º, ..., 12.º)
    def sort_key(g):
        m = re.match(r"(\d+)", g or "")
        return (int(m.group(1)) if m else 999, g or "")
    grades = sorted([g for g in grades if g], key=sort_key)
    return {"grades": grades, "active_year": year}


@api.get("/adoptions/availability")
async def adoptions_availability(school_id: str, grade: Optional[str] = None):
    """Indica se existe lista oficial para uma escola/ano no ano ativo.
    Não altera dados; serve só para UX (mensagem quando não há adoções)."""
    school = await db.schools.find_one({"id": school_id}, {"_id": 0})
    if not school:
        raise HTTPException(404, "Escola não encontrada")

    mun = await db.municipalities.find_one({"id": school.get("municipality_id")}, {"_id": 0, "name": 1})
    year = await _get_active_school_year()
    if not mun or not year:
        return {
            "has_adoptions": False,
            "count": 0,
            "active_year": year,
            "concelho": mun.get("name") if mun else None,
            "escola": school.get("name"),
        }

    filt: Dict[str, Any] = {
        "school_year": year,
        "concelho": mun.get("name"),
        "escola": school.get("name"),
    }
    if grade:
        filt["grade"] = grade
    count = await db.school_adoptions.count_documents(filt)
    return {
        "has_adoptions": count > 0,
        "count": count,
        "active_year": year,
        "concelho": mun.get("name"),
        "escola": school.get("name"),
    }


@api.get("/adoptions/books")
async def adoptions_books(concelho: str, escola: str, grade: str):
    """Devolve os manuais adotados para uma escola/ano, com dados do catálogo
    quando existem (preço, ID) ou apenas título quando não estão."""
    year = await _get_active_school_year()
    if not year:
        raise HTTPException(404, "Não há adoções ativas.")
    cursor = db.school_adoptions.find(
        {"school_year": year, "concelho": concelho, "escola": escola, "grade": grade},
        {"_id": 0},
    ).sort("subject", 1)
    adoptions: List[dict] = []
    isbns: set = set()
    async for a in cursor:
        adoptions.append(a)
        isbns.add(a["isbn13"])
    # Enriquecer com dados do catálogo
    catalog = {}
    if isbns:
        async for b in db.books.find({"isbn13": {"$in": list(isbns)}}, {"_id": 0}):
            catalog[b["isbn13"]] = b
    result = []
    for a in adoptions:
        b = catalog.get(a["isbn13"])
        result.append({
            "isbn13": a["isbn13"],
            "subject": a.get("subject") or "Sem disciplina",
            "title": (b.get("title") if b else a.get("title")) or "",
            "publisher": (b.get("publisher") if b else a.get("publisher")) or "",
            "in_catalog": bool(b),
            "price": float(b.get("price", 0)) if b else None,
            "slug": b.get("slug") if b else None,
            "image_url": b.get("image_url") if b else None,
            "status": b.get("status") if b else None,
        })
    return {
        "school_year": year,
        "concelho": concelho,
        "escola": escola,
        "grade": grade,
        "books": result,
    }


@api.post("/admin/schools/import")
async def import_schools(file: UploadFile = File(...), admin: dict = Depends(require_super_admin)):
    """Importa escolas e anos a partir de um ficheiro Excel.
    Espera as colunas Escola, Município e Anos."""
    content = await file.read()
    try:
        df = pd.read_excel(BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Falha ao ler ficheiro Excel: {e}")

    df.columns = [str(c).strip().lower() for c in df.columns]
    col_map = {
        "escola": "school", "school": "school",
        "concelho": "municipality",
        "município": "municipality", "municipio": "municipality", "municipality": "municipality",
        "anos": "grades", "grades": "grades",
    }
    df.rename(columns={k: v for k, v in col_map.items() if k in df.columns}, inplace=True)

    mun_cache: Dict[str, str] = {}
    school_cache: Dict[tuple, dict] = {}
    created_schools = updated_schools = created_munis = anomalies = 0
    issues = []

    for _, row in df.iterrows():
        school_name = str(row.get("school", "")).strip()
        mun_name = str(row.get("municipality", "")).strip()
        grades = _parse_grade_list(str(row.get("grades", "")).strip())
        if not school_name or not mun_name or not grades:
            anomalies += 1
            issues.append({"school": school_name, "municipality": mun_name, "grades": row.get("grades", ""), "issue": "Escola, Município ou Anos em falta/inválidos"})
            continue

        mun_key = mun_name.lower()
        if mun_key not in mun_cache:
            mun_doc = await db.municipalities.find_one({"name": {"$regex": f"^{re.escape(mun_name)}$", "$options": "i"}})
            if not mun_doc:
                mun_doc = {"id": gen_id(), "name": mun_name}
                await db.municipalities.insert_one(mun_doc)
                created_munis += 1
            mun_cache[mun_key] = mun_doc["id"]
        mun_id = mun_cache[mun_key]

        school_key = (school_name.lower(), mun_id)
        if school_key not in school_cache:
            school_doc = await db.schools.find_one({"name": {"$regex": f"^{re.escape(school_name)}$", "$options": "i"}, "municipality_id": mun_id})
            if school_doc:
                school_cache[school_key] = school_doc
        school_doc = school_cache.get(school_key)

        if school_doc:
            school_doc_grades = school_doc.get("grades_taught") or []
            if school_doc_grades != grades or school_doc.get("name") != school_name:
                await db.schools.update_one({"id": school_doc["id"]}, {"$set": {"name": school_name, "grades_taught": grades}})
                updated_schools += 1
            school_cache[school_key] = {**school_doc, "name": school_name, "grades_taught": grades}
        else:
            school_doc = {"id": gen_id(), "name": school_name, "municipality_id": mun_id, "grades_taught": grades}
            await db.schools.insert_one(school_doc)
            created_schools += 1
            school_cache[school_key] = school_doc

    await log_action(admin["id"], "import", "schools", None, {
        "created_schools": created_schools, "updated_schools": updated_schools,
        "created_municipalities": created_munis, "anomalies": anomalies,
    })
    return {
        "created_schools": created_schools,
        "updated_schools": updated_schools,
        "created_municipalities": created_munis,
        "anomalies": anomalies,
        "issues": issues[:50],
    }

@api.post("/admin/books/import")
async def import_books(file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    """Legacy one-shot import kept for backward compatibility. Now also
    enforces the 'never import invalid rows' rule (price <= 0 are skipped)."""
    content = await file.read()
    df = _normalize_import_df(content)
    classified = _classify_import_rows(df)

    created = updated = anomalies = 0
    issues = []
    for r in classified:
        if r["action"] == "error":
            anomalies += 1
            issues.append({"isbn": r["isbn"], "title": r["title"], "issue": r["issue"]})
            continue
        data = r["data"]
        isbn = data["isbn13"]
        existing = await db.books.find_one({"isbn13": isbn})
        if existing:
            await db.books.update_one(
                {"isbn13": isbn},
                {"$set": {
                    "price": data["price"], "type": data["type"],
                    "title": data["title"] or existing.get("title", ""),
                    "publisher": data["publisher"] or existing.get("publisher", ""),
                    "author": data["author"] or existing.get("author", ""),
                    "subject": data["subject"] or existing.get("subject", ""),
                    "updated_at": iso(now_utc()),
                }},
            )
            updated += 1
        else:
            await db.books.insert_one({
                "id": gen_id(), "isbn13": isbn, "title": data["title"],
                "author": data["author"], "publisher": data["publisher"],
                "subject": data["subject"], "price": data["price"], "type": data["type"],
                "status": "Available", "stock_qty": 0, "synopsis": "",
                "features": {"cycle": data["cycle"], "grade": data["grade_level"]},
                "image_url": "", "is_lamination_eligible": True,
                "created_at": iso(now_utc()),
            })
            created += 1
    await log_action(admin["id"], "import", "books", None, {"created": created, "updated": updated, "anomalies": anomalies})
    return {"created": created, "updated": updated, "anomalies": anomalies, "issues": issues[:50]}

# Enrich missing book covers via Google Books API
# Google Books own-quota key (avoids the shared-IP 429s). Set in backend/.env.
GOOGLE_BOOKS_API_KEY = os.environ.get("GOOGLE_BOOKS_API_KEY", "").strip()


class ISBNdbAuthError(Exception):
    pass


class ISBNdbRateLimitError(Exception):
    def __init__(self, retry_after: Optional[str] = None):
        super().__init__("ISBNdb rate limited")
        self.retry_after = retry_after


def _gb_url(query: str) -> str:
    """Google Books API URL builder. `country=PT` is REQUIRED (without it the
    API returns 403 'unknownLocation' because it can't derive the caller's
    location from a data-center IP). The `key=` parameter is appended when a
    dedicated key is configured, moving us off the shared free-tier quota."""
    url = f"https://www.googleapis.com/books/v1/volumes?{query}&country=PT"
    if GOOGLE_BOOKS_API_KEY:
        url += f"&key={GOOGLE_BOOKS_API_KEY}"
    return url


async def _image_ok(client_http, url: str) -> bool:
    try:
        r = await client_http.get(url, follow_redirects=True)
        ctype = r.headers.get("content-type", "")
        return r.status_code == 200 and ctype.startswith("image") and len(r.content) > 1500
    except Exception:
        return False


async def _throttle_isbndb_request():
    global _ISBNDB_LAST_REQUEST_AT
    async with _ISBNDB_LOCK:
        now = time.monotonic()
        wait_for = (_ISBNDB_LAST_REQUEST_AT + ISBNDB_MIN_INTERVAL) - now
        if wait_for > 0:
            await asyncio.sleep(wait_for)
            now = time.monotonic()
        _ISBNDB_LAST_REQUEST_AT = now


async def _resolve_isbndb_cover(client_http, isbn: str) -> Optional[str]:
    clean_isbn = strip_isbn(isbn or "")
    if not ISBNDB_API_KEY or not re.fullmatch(r"\d{13}", clean_isbn):
        return None

    await _throttle_isbndb_request()
    url = f"{ISBNDB_BASE_URL}/book/{clean_isbn}?with_prices=false"
    headers = {"Authorization": ISBNDB_API_KEY, "Accept": "application/json"}

    try:
        r = await client_http.get(url, headers=headers)
    except Exception as e:
        logger.warning(f"[covers] ISBNdb exception for ISBN {clean_isbn}: {type(e).__name__}: {e}")
        return None

    if r.status_code in (401, 403):
        logger.warning("[covers] ISBNdb auth error. Verifique ISBNDB_API_KEY.")
        raise ISBNdbAuthError("ISBNdb recusou a autenticação. Verifique ISBNDB_API_KEY.")
    if r.status_code == 429:
        retry_after = r.headers.get("retry-after")
        logger.warning("[covers] ISBNdb rate limit atingido.")
        raise ISBNdbRateLimitError(retry_after)
    if r.status_code == 404:
        logger.info(f"[covers] ISBNdb 404 para ISBN {clean_isbn}")
        return None
    if r.status_code >= 400:
        logger.warning(f"[covers] ISBNdb HTTP {r.status_code} para ISBN {clean_isbn}: {r.text[:200]}")
        return None

    try:
        data = r.json()
    except Exception as e:
        logger.warning(f"[covers] ISBNdb JSON inválido para ISBN {clean_isbn}: {type(e).__name__}: {e}")
        return None

    image = (data.get("book") or {}).get("image")
    if image and await _image_ok(client_http, image):
        return image.replace("http://", "https://")

    logger.info(f"[covers] ISBNdb sem imagem válida para ISBN {clean_isbn}")
    return None


async def _resolve_cover_url(
    client_http,
    book: dict,
    publisher_template: Optional[str] = None,
    diag: Optional[dict] = None,
) -> Optional[dict]:
    """Tries several public sources, in order of reliability, and returns the
    first WORKING cover URL (or None). Records per-source outcomes into
    `diag` (dict keyed by source name → {success,not_found,blocked,error}) so
    the admin UI can show WHY a run produced few/no covers.

    Sources tried:
      0. Publisher URL template (configured in Settings) — validated before use.
      1. ISBNdb by ISBN (when configured, with 1 req/s throttle)
      2. Google Books by ISBN (with own API key + country=PT)
      3. Google Books by title (+ author when available)
      4. Open Library by ISBN
    """
    isbn = book.get("isbn13", "") or ""
    title = (book.get("title", "") or "").strip()
    author = (book.get("author", "") or "").strip()
    isbn_is_valid = bool(re.fullmatch(r"\d{13}", strip_isbn(isbn)))

    def note(source: str, outcome: str):
        if diag is None:
            return
        d = diag.setdefault(source, {"success": 0, "not_found": 0, "blocked": 0, "error": 0})
        d[outcome] = d.get(outcome, 0) + 1

    # 0. Publisher cover template — TRUSTED (see docstring).
    if publisher_template and "{isbn}" in publisher_template:
        pub_url = publisher_template.replace("{isbn}", isbn)
        if await _image_ok(client_http, pub_url):
            note("publisher_template", "success")
            return {"url": pub_url, "source": "publisher"}
        note("publisher_template", "not_found")

    # 1. ISBNdb by ISBN
    if isbn_is_valid and ISBNDB_API_KEY:
        try:
            isbndb_url = await _resolve_isbndb_cover(client_http, isbn)
            if isbndb_url:
                note("isbndb", "success")
                return {"url": isbndb_url, "source": "isbndb"}
            note("isbndb", "not_found")
        except ISBNdbAuthError:
            note("isbndb", "error")
            raise
        except ISBNdbRateLimitError:
            note("isbndb", "blocked")
            raise
        except Exception as e:
            note("isbndb", "error")
            logger.warning(f"[covers] ISBNdb unexpected exception for {isbn}: {type(e).__name__}: {e}")
    elif not isbn_is_valid:
        note("isbndb", "not_found")

    # 2. Google Books by ISBN
    if isbn_is_valid:
        try:
            r = await client_http.get(_gb_url(f"q=isbn:{isbn}"))
            if r.status_code == 429:
                note("google_isbn", "blocked")
                logger.warning(f"[covers] Google Books quota (429) for ISBN {isbn}. Set GOOGLE_BOOKS_API_KEY.")
            elif r.status_code >= 400:
                note("google_isbn", "error")
                logger.warning(f"[covers] Google Books HTTP {r.status_code} for ISBN {isbn}: {r.text[:200]}")
            else:
                data = r.json()
                if data.get("totalItems", 0) > 0:
                    links = (data["items"][0].get("volumeInfo", {}).get("imageLinks") or {})
                    u = links.get("thumbnail") or links.get("smallThumbnail")
                    if u:
                        note("google_isbn", "success")
                        return {"url": u.replace("http://", "https://").replace("&edge=curl", ""), "source": "google_books"}
                    note("google_isbn", "not_found")
                else:
                    note("google_isbn", "not_found")
        except Exception as e:
            note("google_isbn", "error")
            logger.warning(f"[covers] Google Books ISBN exception for {isbn}: {type(e).__name__}: {e}")

    # 3. Google Books by title (+ author when present)
    if title:
        try:
            q = f'intitle:"{title}"'
            if author:
                q += f'+inauthor:"{author}"'
            r = await client_http.get(_gb_url(f"q={q}&maxResults=1"))
            if r.status_code == 429:
                note("google_title", "blocked")
            elif r.status_code >= 400:
                note("google_title", "error")
                logger.warning(f"[covers] Google Books title HTTP {r.status_code} for ISBN {isbn}: {r.text[:200]}")
            else:
                data = r.json()
                if data.get("totalItems", 0) > 0:
                    links = (data["items"][0].get("volumeInfo", {}).get("imageLinks") or {})
                    u = links.get("thumbnail") or links.get("smallThumbnail")
                    if u:
                        note("google_title", "success")
                        return {"url": u.replace("http://", "https://").replace("&edge=curl", ""), "source": "google_books"}
                    note("google_title", "not_found")
                else:
                    note("google_title", "not_found")
        except Exception as e:
            note("google_title", "error")
            logger.warning(f"[covers] Google Books title exception for {isbn}: {type(e).__name__}: {e}")

    # 4. Open Library by ISBN
    if isbn_is_valid:
        ol_url = f"https://covers.openlibrary.org/b/isbn/{isbn}-L.jpg?default=false"
        try:
            r = await client_http.get(ol_url, follow_redirects=True)
            ctype = r.headers.get("content-type", "")
            if r.status_code == 200 and ctype.startswith("image") and len(r.content) > 1500:
                note("openlibrary", "success")
                return {"url": ol_url, "source": "open_library"}
            elif r.status_code == 404:
                note("openlibrary", "not_found")
            else:
                note("openlibrary", "error")
        except Exception as e:
            note("openlibrary", "error")
            logger.warning(f"[covers] Open Library exception for {isbn}: {type(e).__name__}: {e}")

    return {"url": None, "source": None, "skipped_no_isbn": not isbn_is_valid}


def _missing_cover_query() -> Dict[str, Any]:
    return {
        "$or": [
            {"image_url": ""},
            {"image_url": None},
            {
                "$and": [
                    {"image_url": {"$regex": "openlibrary"}},
                    {"cover_source": {"$ne": "open_library"}},
                ]
            },
        ]
    }


@api.get("/admin/books/covers-status")
async def covers_status(admin: dict = Depends(require_admin)):
    """How many books still need a real cover (so the UI can show progress)."""
    total = await db.books.count_documents({})
    missing = await db.books.count_documents(_missing_cover_query())
    return {"total": total, "with_cover": total - missing, "missing": missing}


@api.post("/admin/books/enrich-covers")
async def enrich_covers(admin: dict = Depends(require_admin), limit: int = 50, cursor: Optional[str] = None):
    """Fills in missing book covers from public sources (see _resolve_cover_url).
    Processes up to `limit` books per call so a large catalog can be done in
    batches without timing out. Returns:
      - updated / processed / remaining / done / next_cursor
      - source counters and rate_limit flags so the admin can see progress.
    """
    import httpx
    updated = 0
    processed = 0
    not_found = 0
    skipped_no_isbn = 0
    rate_limited = False
    next_cursor = cursor
    source_counts = {"publisher": 0, "isbndb": 0, "google_books": 0, "open_library": 0}
    settings = await db.settings.find_one({"id": "global"}, {"_id": 0}) or {}
    publisher_template = settings.get("publisher_cover_template") or None
    diagnostics: dict = {}
    missing_filter: Dict[str, Any] = _missing_cover_query()
    query: Dict[str, Any] = dict(missing_filter)
    if cursor:
        query["id"] = {"$gt": cursor}
    batch = await db.books.find(query, {"_id": 0}).sort("id", 1).limit(limit).to_list(limit)
    async with httpx.AsyncClient(timeout=10.0, headers={"User-Agent": "TendinhaDoSaber/1.0"}) as client_http:
        for b in batch:
            processed += 1
            next_cursor = b.get("id") or next_cursor
            try:
                resolved = await _resolve_cover_url(client_http, b, publisher_template, diag=diagnostics)
                url = (resolved or {}).get("url")
                source = (resolved or {}).get("source")
                if url and source:
                    await db.books.update_one({"id": b["id"]}, {"$set": {"image_url": url, "cover_source": source}})
                    updated += 1
                    if source in source_counts:
                        source_counts[source] += 1
                else:
                    not_found += 1
                    if (resolved or {}).get("skipped_no_isbn"):
                        skipped_no_isbn += 1
            except ISBNdbRateLimitError:
                rate_limited = True
                logger.warning("[covers] ISBNdb rate limit atingido. Parando lote atual.")
                break
            except ISBNdbAuthError:
                logger.warning("[covers] ISBNdb recusou a autenticação. Parando lote atual.")
                raise HTTPException(status_code=403, detail="ISBNdb recusou a autenticação. Verifique ISBNDB_API_KEY.")
            except Exception as e:
                logger.warning(f"[covers] Unexpected exception for {b.get('isbn13')}: {type(e).__name__}: {e}")
                diagnostics.setdefault("_unhandled", {"count": 0})
                diagnostics["_unhandled"]["count"] += 1
                continue
    remaining_query: Dict[str, Any] = dict(missing_filter)
    if next_cursor:
        remaining_query["id"] = {"$gt": next_cursor}
    remaining = await db.books.count_documents(remaining_query)
    done = remaining == 0 and not rate_limited
    await log_action(admin["id"], "enrich", "covers", None, {"updated": updated, "processed": processed, "diag": diagnostics, "rate_limited": rate_limited, "next_cursor": next_cursor})
    return {
        "updated": updated,
        "processed": processed,
        "remaining": remaining,
        "done": done,
        "next_cursor": next_cursor,
        "sources": source_counts,
        "not_found": not_found,
        "skipped_no_isbn": skipped_no_isbn,
        "rate_limited": rate_limited,
        "diagnostics": diagnostics,
        "api_key_configured": bool(GOOGLE_BOOKS_API_KEY),
        "publisher_template_configured": bool(publisher_template and "{isbn}" in (publisher_template or "")),
    }

# ---------- Municipalities / Schools ----------
@api.get("/municipalities")
async def list_municipalities():
    return await db.municipalities.find({}, {"_id": 0}).sort("name", 1).to_list(500)

@api.get("/schools")
async def list_schools(municipality_id: Optional[str] = None, grade: Optional[str] = None):
    filt: Dict[str, Any] = {}
    if municipality_id:
        filt["municipality_id"] = municipality_id
    if grade:
        filt["grades_taught"] = grade
    return await db.schools.find(filt, {"_id": 0}).sort("name", 1).to_list(2000)

@api.get("/grade-levels")
async def list_grades():
    return list(_GRADES_ALL)

@api.post("/admin/municipalities")
async def create_mun(payload: MunicipalityIn, admin: dict = Depends(require_admin)):
    doc = {"id": gen_id(), "name": payload.name}
    await db.municipalities.insert_one(doc)
    await log_action(admin["id"], "create", "municipality", doc["id"])
    doc.pop("_id", None)
    return doc

@api.delete("/admin/municipalities/{mid}")
async def delete_mun(mid: str, admin: dict = Depends(require_admin)):
    await db.municipalities.delete_one({"id": mid})
    await db.schools.delete_many({"municipality_id": mid})
    await log_action(admin["id"], "delete", "municipality", mid)
    return {"ok": True}

@api.post("/admin/schools")
async def create_school(payload: SchoolIn, admin: dict = Depends(require_super_admin)):
    doc = {
        "id": gen_id(),
        "name": payload.name,
        "municipality_id": payload.municipality_id,
        "grades_taught": payload.grades_taught if payload.grades_taught is not None else _GRADES_ALL,
    }
    await db.schools.insert_one(doc)
    await log_action(admin["id"], "create", "school", doc["id"])
    doc.pop("_id", None)
    return doc

@api.put("/admin/schools/{sid}")
async def update_school(sid: str, payload: SchoolIn, admin: dict = Depends(require_super_admin)):
    update = {
        "name": payload.name,
        "municipality_id": payload.municipality_id,
        "grades_taught": payload.grades_taught if payload.grades_taught is not None else _GRADES_ALL,
    }
    res = await db.schools.update_one({"id": sid}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(404, "Escola não encontrada")
    await log_action(admin["id"], "update", "school", sid)
    return clean_doc(await db.schools.find_one({"id": sid}))

@api.delete("/admin/schools/{sid}")
async def delete_school(sid: str, admin: dict = Depends(require_super_admin)):
    await db.schools.delete_one({"id": sid})
    await db.school_books.delete_many({"school_id": sid})
    await log_action(admin["id"], "delete", "school", sid)
    return {"ok": True}

@api.delete("/admin/schools")
async def delete_all_schools(
    confirmation: str = Form(...),
    admin: dict = Depends(require_manager),
):
    """DANGER: wipes every document from `schools` AND removes the orphan
    `school_books` links that would otherwise dangle. Restricted to
    admin/super_admin. Requires exact confirmation phrase 'APAGAR TODAS'.
    Counts before/after are returned so the UI can show what happened."""
    if (confirmation or "").strip() != "APAGAR TODAS":
        raise HTTPException(400, "Confirmação inválida. Escreva exatamente 'APAGAR TODAS' para confirmar.")
    schools_before = await db.schools.count_documents({})
    links_before = await db.school_books.count_documents({})
    s_res = await db.schools.delete_many({})
    l_res = await db.school_books.delete_many({})
    await log_action(
        admin["id"], "bulk_delete", "schools", "ALL",
        {"schools_deleted": s_res.deleted_count, "links_deleted": l_res.deleted_count},
    )
    logger.warning(
        f"[ADMIN BULK WIPE] {admin.get('email')} apagou {s_res.deleted_count} escolas "
        f"e {l_res.deleted_count} ligações school_books."
    )
    return {
        "ok": True,
        "schools_before": schools_before,
        "schools_deleted": s_res.deleted_count,
        "school_books_before": links_before,
        "school_books_deleted": l_res.deleted_count,
    }

@api.post("/admin/school-books")
async def link_school_book(payload: SchoolBookIn, admin: dict = Depends(require_admin)):
    payload_isbn = strip_isbn(payload.isbn13)
    existing = await db.school_books.find_one({"school_id": payload.school_id, "isbn13": payload_isbn, "grade_level": payload.grade_level})
    if existing:
        return clean_doc(existing)
    doc = {"id": gen_id(), "school_id": payload.school_id, "isbn13": payload_isbn, "grade_level": payload.grade_level}
    await db.school_books.insert_one(doc)
    await log_action(admin["id"], "link", "school_book", doc["id"])
    doc.pop("_id", None)
    return doc

@api.delete("/admin/school-books/{link_id}")
async def unlink_school_book(link_id: str, admin: dict = Depends(require_admin)):
    await db.school_books.delete_one({"id": link_id})
    return {"ok": True}

@api.post("/admin/school-books/import")
async def import_school_books(file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    """Imports the Escola<->ISBN relationships file: columns ISBN | Município
    (Municipality) | Escola (School) | Disciplina (Subject) | Ano (Grade
    Level). Municípios and Escolas are found-or-created by name so this can
    run standalone even before any school exists yet."""
    content = await file.read()
    try:
        df = pd.read_excel(BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Falha ao ler ficheiro Excel: {e}")

    df.columns = [str(c).strip().lower() for c in df.columns]
    col_map = {
        "isbn": "isbn13", "município": "municipality", "municipio": "municipality",
        "municipality": "municipality", "escola": "school", "school": "school",
        "disciplina": "subject", "subject": "subject",
        "ano": "grade_level", "ano de escolaridade": "grade_level", "grade level": "grade_level", "grade_level": "grade_level",
    }
    df.rename(columns={k: v for k, v in col_map.items() if k in df.columns}, inplace=True)

    mun_cache: Dict[str, str] = {}
    school_cache: Dict[tuple, str] = {}
    created_links = updated_links = created_schools = created_munis = anomalies = 0
    issues = []

    for _, row in df.iterrows():
        isbn = strip_isbn(str(row.get("isbn13", "")))
        mun_name = str(row.get("municipality", "")).strip()
        school_name = str(row.get("school", "")).strip()
        grade = _grade_to_label(row.get("grade_level", ""))
        if len(isbn) != 13 or not school_name or not mun_name or not grade:
            anomalies += 1
            issues.append({"isbn": isbn, "school": school_name, "issue": "ISBN/Município/Escola/Ano em falta ou inválido"})
            continue

        mun_key = mun_name.lower()
        if mun_key not in mun_cache:
            mun_doc = await db.municipalities.find_one({"name": {"$regex": f"^{re.escape(mun_name)}$", "$options": "i"}})
            if not mun_doc:
                mun_doc = {"id": gen_id(), "name": mun_name}
                await db.municipalities.insert_one(mun_doc)
                created_munis += 1
            mun_cache[mun_key] = mun_doc["id"]
        mun_id = mun_cache[mun_key]

        school_key = (school_name.lower(), mun_id)
        if school_key not in school_cache:
            school_doc = await db.schools.find_one({"name": {"$regex": f"^{re.escape(school_name)}$", "$options": "i"}, "municipality_id": mun_id})
            if not school_doc:
                school_doc = {"id": gen_id(), "name": school_name, "municipality_id": mun_id, "grades_taught": []}
                await db.schools.insert_one(school_doc)
                created_schools += 1
            school_cache[school_key] = school_doc["id"]
        school_id = school_cache[school_key]

        existing = await db.school_books.find_one({"school_id": school_id, "isbn13": isbn, "grade_level": grade})
        if existing:
            updated_links += 1
            continue
        await db.school_books.insert_one({
            "id": gen_id(), "school_id": school_id, "isbn13": isbn, "grade_level": grade,
            "subject": str(row.get("subject", "")).strip(),
        })
        await db.schools.update_one({"id": school_id}, {"$addToSet": {"grades_taught": grade}})
        created_links += 1

    # Recalculate grades_taught for any school that may still have stale _GRADES_ALL values.
    async for stale_school in db.schools.find({"grades_taught": {"$eq": _GRADES_ALL}}, {"id": 1}):
        school_id = stale_school["id"]
        distinct_grades = await db.school_books.distinct("grade_level", {"school_id": school_id})
        await db.schools.update_one({"id": school_id}, {"$set": {"grades_taught": distinct_grades or []}})

    await log_action(admin["id"], "import", "school_books", None, {
        "created_links": created_links, "created_schools": created_schools,
        "created_municipalities": created_munis, "anomalies": anomalies,
    })
    return {
        "created_links": created_links, "already_existing_links": updated_links,
        "created_schools": created_schools, "created_municipalities": created_munis,
        "anomalies": anomalies, "issues": issues[:50],
    }

# ---------- Partners ----------
@api.get("/partners")
async def list_partners():
    return await db.partners.find({"active": {"$ne": False}}, {"_id": 0}).sort("order", 1).to_list(200)

@api.get("/admin/partners")
async def admin_list_partners(admin: dict = Depends(require_admin)):
    return await db.partners.find({}, {"_id": 0}).sort("order", 1).to_list(200)

@api.post("/admin/partners")
async def create_partner(payload: PartnerIn, admin: dict = Depends(require_manager)):
    doc = payload.model_dump()
    doc["id"] = gen_id()
    doc["promo_code"] = doc["promo_code"].upper()
    doc["usage_count"] = 0
    doc["order"] = await db.partners.count_documents({})
    if await db.partners.find_one({"promo_code": doc["promo_code"]}):
        raise HTTPException(400, "Código promocional já existe")
    await db.partners.insert_one(doc)
    await log_action(admin["id"], "create", "partner", doc["id"])
    doc.pop("_id", None)
    return doc

@api.put("/admin/partners/{pid}")
async def update_partner(pid: str, payload: PartnerIn, admin: dict = Depends(require_manager)):
    update = payload.model_dump()
    update["promo_code"] = update["promo_code"].upper()
    dup = await db.partners.find_one({"promo_code": update["promo_code"], "id": {"$ne": pid}})
    if dup:
        raise HTTPException(400, "Código promocional já está a ser usado por outro parceiro")
    res = await db.partners.update_one({"id": pid}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(404, "Parceiro não encontrado")
    await log_action(admin["id"], "update", "partner", pid)
    return clean_doc(await db.partners.find_one({"id": pid}))

@api.put("/admin/partners/{pid}/reorder")
async def reorder_partner(pid: str, order: int = Form(...), admin: dict = Depends(require_manager)):
    await db.partners.update_one({"id": pid}, {"$set": {"order": order}})
    return {"ok": True}

@api.delete("/admin/partners/{pid}")
async def delete_partner(pid: str, admin: dict = Depends(require_manager)):
    await db.partners.delete_one({"id": pid})
    await log_action(admin["id"], "delete", "partner", pid)
    return {"ok": True}

def _promo_is_valid(promo: dict) -> bool:
    if not promo or promo.get("active") is False:
        return False
    now = now_utc()
    if promo.get("valid_from"):
        try:
            if now < datetime.fromisoformat(promo["valid_from"]).replace(tzinfo=timezone.utc):
                return False
        except Exception:
            pass
    if promo.get("valid_until"):
        try:
            if now > datetime.fromisoformat(promo["valid_until"]).replace(tzinfo=timezone.utc):
                return False
        except Exception:
            pass
    limit = promo.get("usage_limit")
    if limit is not None and promo.get("usage_count", 0) >= limit:
        return False
    return True

# ---------- Categories (future-proofing: mochilas, calculadoras, etc.) ----------
@api.get("/categories")
async def list_categories():
    return await db.categories.find({"is_active": True}, {"_id": 0}).sort("name", 1).to_list(200)

@api.get("/admin/categories")
async def admin_list_categories(admin: dict = Depends(require_admin)):
    return await db.categories.find({}, {"_id": 0}).sort("name", 1).to_list(200)

@api.post("/admin/categories")
async def create_category(payload: CategoryIn, admin: dict = Depends(require_admin)):
    if await db.categories.find_one({"name": payload.name}):
        raise HTTPException(400, "Categoria já existe")
    doc = {"id": gen_id(), **payload.model_dump()}
    await db.categories.insert_one(doc)
    await log_action(admin["id"], "create", "category", doc["id"])
    doc.pop("_id", None)
    return doc

@api.put("/admin/categories/{cid}")
async def update_category(cid: str, payload: CategoryIn, admin: dict = Depends(require_admin)):
    res = await db.categories.update_one({"id": cid}, {"$set": payload.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(404, "Categoria não encontrada")
    await log_action(admin["id"], "update", "category", cid)
    return {"ok": True}

@api.delete("/admin/categories/{cid}")
async def delete_category(cid: str, admin: dict = Depends(require_admin)):
    await db.categories.delete_one({"id": cid})
    await log_action(admin["id"], "delete", "category", cid)
    return {"ok": True}

# ---------- Content / CMS ----------
_CONTENT_DEFAULTS = {
    "about_us": (
        "A Tendinha do Saber nasceu em Aveiro com um objetivo simples: tornar a compra de manuais "
        "escolares e cadernos de atividades mais fácil, rápida e próxima das famílias. Trabalhamos "
        "diretamente com as escolas do concelho para que cada encomenda corresponda exatamente à "
        "lista do ano e da turma do seu filho — sem confusões, sem stress de início de ano letivo."
    ),
    "hero_title": "Os manuais escolares de que precisa, de forma simples, rápida e com a confiança de uma livraria de proximidade.",
    "hero_subtitle": "Encontre os manuais e cadernos de atividades da sua escola, utilize o Voucher MEGA e escolha a plastificação dos seus livros.",
    "footer_text": "Mais do que uma livraria, um parceiro das famílias na escolha dos manuais escolares.",
    "instagram_handle": "@tendinhadosaber",
    "instagram_url": "https://instagram.com/tendinhadosaber",
    "partners_cta": "Tem interesse em tornar-se parceiro da Tendinha do Saber?",
    "promotions_label": "Desconto exclusivo para parceiros",
}

_DEFAULT_FAQS = [
    {
        "question": "Como faço para encomendar os livros da escola do meu filho?",
        "answer": "Use o seletor da página inicial: escolha o ano, o concelho e a escola. Mostramos imediatamente a lista oficial e pode adicionar tudo ao carrinho. Em alternativa, pesquise por título ou ISBN.",
    },
    {
        "question": "Os preços são iguais aos das outras livrarias?",
        "answer": "Sim. Praticamos o Preço de Venda ao Público (PVP) recomendado pelas editoras. A diferença está no serviço de proximidade e na plastificação.",
    },
    {
        "question": "Quanto custa a plastificação?",
        "answer": "2€ por livro. É opcional e escolhe por cada livro no carrinho. Apenas se aplica a manuais (cadernos de fichas não são plastificados).",
    },
    {
        "question": "Tenho um código de parceiro. Onde o aplico?",
        "answer": "Na página do carrinho existe um campo para inserir o código promocional. Basta introduzi-lo para que o desconto seja aplicado automaticamente à sua encomenda.",
    },
    {
        "question": "Vocês entregam em casa?",
        "answer": "Sim. Fazemos entrega ao domicílio no distrito de Aveiro. O custo é apresentado no checkout e depende do concelho selecionado.",
    },
    {
        "question": "Como funciona o voucher MEGA?",
        "answer": "Pode submeter o voucher (código ou PDF) na página dedicada. A nossa equipa valida em 24h úteis e o desconto é aplicado à sua próxima encomenda. Veja o passo-a-passo na página Como funciona o voucher MEGA.",
    },
    {
        "question": "E se o livro estiver indicado como 'Disponível por encomenda'?",
        "answer": "Significa que o artigo não está em stock imediato e será solicitado ao fornecedor. O prazo depende da disponibilidade do fornecedor. Salvo indicação ou acordo diferente, a encomenda será cumprida no prazo legal aplicável, em regra até 30 dias. Se tal não for possível, será informado.",
    },
    {
        "question": "Posso pagar com MB Way?",
        "answer": "Após a sua encomenda, entraremos em contacto para combinar o pagamento pelos meios efetivamente disponíveis nesse momento. A fatura é emitida após confirmação do pagamento.",
    },
]


async def _ensure_default_faqs_seeded():
    if await db.faqs.count_documents({}) > 0:
        return
    now = iso(now_utc())
    docs = [
        {
            "id": gen_id(),
            "question": item["question"],
            "answer": item["answer"],
            "sort_order": idx,
            "created_at": now,
            "updated_at": now,
        }
        for idx, item in enumerate(_DEFAULT_FAQS, start=1)
    ]
    if docs:
        await db.faqs.insert_many(docs)

@api.get("/content")
async def get_content():
    doc = await db.site_content.find_one({"id": "main"}, {"_id": 0}) or {}
    return {**_CONTENT_DEFAULTS, **{k: v for k, v in doc.items() if v not in (None, "") and k != "id"}}

@api.put("/admin/content")
async def update_content(payload: ContentIn, admin: dict = Depends(require_super_admin)):
    update = {k: v for k, v in payload.model_dump().items() if v is not None}
    await db.site_content.update_one({"id": "main"}, {"$set": {**update, "id": "main"}}, upsert=True)
    await log_action(admin["id"], "update", "content", "main", update)
    doc = await db.site_content.find_one({"id": "main"}, {"_id": 0}) or {}
    return {**_CONTENT_DEFAULTS, **{k: v for k, v in doc.items() if v not in (None, "") and k != "id"}}


class FAQIn(BaseModel):
    question: str
    answer: str
    sort_order: Optional[int] = None


class FAQReorderItem(BaseModel):
    id: str
    sort_order: int


class FAQReorderIn(BaseModel):
    items: List[FAQReorderItem]


@api.get("/faq")
async def public_faq_list():
    await _ensure_default_faqs_seeded()
    items = await db.faqs.find({}, {"_id": 0}).sort([("sort_order", 1), ("created_at", 1)]).to_list(200)
    return {"items": items}


@api.get("/admin/faq")
async def admin_faq_list(admin: dict = Depends(require_admin)):
    await _ensure_default_faqs_seeded()
    items = await db.faqs.find({}, {"_id": 0}).sort([("sort_order", 1), ("created_at", 1)]).to_list(200)
    return {"items": items}


@api.post("/admin/faq")
async def admin_faq_create(payload: FAQIn, admin: dict = Depends(require_manager)):
    question = (payload.question or "").strip()
    answer = (payload.answer or "").strip()
    if not question or not answer:
        raise HTTPException(400, "Pergunta e resposta são obrigatórias.")
    max_doc = await db.faqs.find_one({}, sort=[("sort_order", -1)], projection={"_id": 0, "sort_order": 1}) or {}
    sort_order = payload.sort_order if payload.sort_order is not None else int(max_doc.get("sort_order") or 0) + 1
    doc = {
        "id": gen_id(),
        "question": question,
        "answer": answer,
        "sort_order": max(1, int(sort_order)),
        "created_at": iso(now_utc()),
        "updated_at": iso(now_utc()),
    }
    await db.faqs.insert_one(doc)
    await log_action(admin["id"], "create", "faq", doc["id"])
    doc.pop("_id", None)
    return doc


@api.put("/admin/faq/{faq_id}")
async def admin_faq_update(faq_id: str, payload: FAQIn, admin: dict = Depends(require_manager)):
    existing = await db.faqs.find_one({"id": faq_id}, {"_id": 0, "id": 1})
    if not existing:
        raise HTTPException(404, "FAQ não encontrada")
    question = (payload.question or "").strip()
    answer = (payload.answer or "").strip()
    if not question or not answer:
        raise HTTPException(400, "Pergunta e resposta são obrigatórias.")
    update = {
        "question": question,
        "answer": answer,
        "updated_at": iso(now_utc()),
    }
    if payload.sort_order is not None:
        update["sort_order"] = max(1, int(payload.sort_order))
    await db.faqs.update_one({"id": faq_id}, {"$set": update})
    await log_action(admin["id"], "update", "faq", faq_id)
    return {"ok": True}


@api.delete("/admin/faq/{faq_id}")
async def admin_faq_delete(faq_id: str, admin: dict = Depends(require_manager)):
    res = await db.faqs.delete_one({"id": faq_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "FAQ não encontrada")
    await log_action(admin["id"], "delete", "faq", faq_id)
    return {"ok": True}


@api.post("/admin/faq/reorder")
async def admin_faq_reorder(payload: FAQReorderIn, admin: dict = Depends(require_manager)):
    for item in payload.items:
        await db.faqs.update_one(
            {"id": item.id},
            {"$set": {"sort_order": max(1, int(item.sort_order)), "updated_at": iso(now_utc())}},
        )
    await log_action(admin["id"], "reorder", "faq", None, {"count": len(payload.items)})
    return {"ok": True}

# ---------- Cart / Promo ----------
async def _compute_cart(items: List[CartItem], promo_code: Optional[str], bags_qty: int = 0) -> dict:
    promo = None
    if promo_code:
        candidate = await db.partners.find_one({"promo_code": promo_code.upper()}, {"_id": 0})
        if candidate and _promo_is_valid(candidate):
            promo = candidate

    settings = await db.settings.find_one({"id": "global"}, {"_id": 0}) or {}
    lam_price = float(settings.get("lamination_price", LAMINATION_PRICE))

    lines = []
    subtotal_manuals = 0.0
    subtotal_workbooks = 0.0
    discount_workbooks = 0.0
    lamination_total = 0.0

    for it in items:
        # Bloco A: aceita isbn13, slug ou pe_code como identificador do item
        book = await _find_book_by_key(it.isbn13)
        if not book:
            continue
        line_price = float(book["price"]) * it.qty
        line_lam = lam_price * it.qty if (it.lamination and book.get("is_lamination_eligible", True)) else 0.0
        is_workbook = book.get("type") == "Workbook"
        line_discount = 0.0
        if promo and is_workbook:
            line_discount = line_price * (float(promo["discount_value"]) / 100.0)
        if is_workbook:
            subtotal_workbooks += line_price
            discount_workbooks += line_discount
        else:
            subtotal_manuals += line_price
        lamination_total += line_lam
        lines.append({
            "isbn13": book["isbn13"],
            "title": book["title"],
            "image_url": book.get("image_url", ""),
            "qty": it.qty,
            "unit_price": book["price"],
            "type": book.get("type", "Manual"),
            "lamination": bool(line_lam > 0),
            "lamination_total": round(line_lam, 2),
            "line_subtotal": round(line_price, 2),
            "line_discount": round(line_discount, 2),
            "line_total": round(line_price - line_discount + line_lam, 2),
        })

    bags_qty = max(0, int(bags_qty or 0))
    bags_total = round(bags_qty * BAG_PRICE, 2)
    total = subtotal_manuals + subtotal_workbooks - discount_workbooks + lamination_total + bags_total
    return {
        "lines": lines,
        "subtotal_manuals": round(subtotal_manuals, 2),
        "subtotal_workbooks": round(subtotal_workbooks, 2),
        "discount_workbooks": round(discount_workbooks, 2),
        "lamination_total": round(lamination_total, 2),
        "bags_qty": bags_qty,
        "bags_total": bags_total,
        "total": round(total, 2),
        "promo": {"code": promo["promo_code"], "discount_value": promo["discount_value"], "partner": promo["name"]} if promo else None,
        "lamination_price": lam_price,
        "shipping_flat_rate": float(settings.get("shipping_flat_rate", SHIPPING_FLAT_RATE)),
    }

@api.post("/cart/validate")
async def cart_validate(payload: PromoValidateIn):
    return await _compute_cart(payload.items, payload.promo_code, payload.bags_qty or 0)

# ---------- Postcodes (Aveiro geofencing — distrito) ----------
@api.get("/postcode/check")
async def check_postcode(code: str):
    code_clean = code.strip().split("-")[0][:4] if code else ""
    settings = await db.settings.find_one({"id": "global"}, {"_id": 0}) or {}
    aveiro = settings.get("aveiro_postcodes", _AVEIRO_CONCELHO_POSTCODES)
    hand_delivery_available = len(code_clean) == 4 and code_clean[:2] in ("37", "38")
    # Bloco B: indica se o CP pertence ao DISTRITO de Aveiro (mais permissivo do
    # que "concelho de Aveiro"). O checkout usa isto para AVISAR sem bloquear.
    in_aveiro_district = _postcode_in_aveiro_district(code_clean)
    return {
        "hand_delivery_available": hand_delivery_available,
        "in_aveiro_district": in_aveiro_district,
        "postcode": code_clean,
        "valid_zones": aveiro,
    }

# ---------- Stock reservation ----------
async def _reserve_stock(lines: List[dict]) -> List[dict]:
    """Atomically decrements stock for 'Available' lines so the same
    physical copy can't be sold twice. Returns the adjustments actually
    applied, so they can be rolled back (compensating action on a later
    failure, or stock restored if the order is later cancelled)."""
    reserved: List[dict] = []
    for line in lines:
        book = await db.books.find_one({"isbn13": line["isbn13"]}, {"_id": 0})
        if not book or book.get("status") == "Unavailable":
            await _restore_stock(reserved)
            raise HTTPException(400, f"'{line.get('title', line['isbn13'])}' já não está disponível.")
        if book.get("status") == "Available":
            res = await db.books.update_one(
                {"isbn13": line["isbn13"], "stock_qty": {"$gte": line["qty"]}},
                {"$inc": {"stock_qty": -line["qty"]}},
            )
            if res.matched_count == 0:
                await _restore_stock(reserved)
                raise HTTPException(409, f"Já não há stock suficiente de '{book.get('title')}'.")
            reserved.append({"isbn13": line["isbn13"], "qty": line["qty"]})
        # status == PreOrder: nothing physical to reserve, always allowed.
    return reserved

async def _restore_stock(adjustments: List[dict]):
    for adj in adjustments:
        await db.books.update_one({"isbn13": adj["isbn13"]}, {"$inc": {"stock_qty": adj["qty"]}})

# Bloco B: Concelhos do distrito de Aveiro (canónicos, ordem alfabética).
# Esta lista é a fonte-de-verdade para o dropdown do checkout e para o
# painel de custos de entrega do admin. Se um concelho não estiver aqui,
# NÃO é aceite como opção de entrega em mão.
AVEIRO_CONCELHOS = [
    "Águeda", "Albergaria-a-Velha", "Anadia", "Arouca", "Aveiro",
    "Castelo de Paiva", "Espinho", "Estarreja", "Ílhavo", "Mealhada",
    "Murtosa", "Oliveira de Azeméis", "Oliveira do Bairro", "Ovar",
    "Santa Maria da Feira", "São João da Madeira", "Sever do Vouga",
    "Vagos", "Vale de Cambra",
]


def _default_shipping_rates() -> Dict[str, float]:
    """Bloco B: valor por defeito para cada concelho é 0 € (grátis).
    O admin pode ajustar depois em /admin/entregas."""
    return {c: 0.0 for c in AVEIRO_CONCELHOS}


async def _get_shipping_rates() -> Dict[str, float]:
    """Devolve o mapa concelho→preço a partir das settings, aplicando o default
    para concelhos que ainda não estejam guardados (para o admin poder ver
    todos os 19 mesmo antes de os configurar pela 1.ª vez)."""
    settings = await db.settings.find_one({"id": "global"}, {"_id": 0}) or {}
    stored = settings.get("shipping_rates") or {}
    rates = _default_shipping_rates()
    for k, v in stored.items():
        if k in rates:
            try:
                rates[k] = round(float(v), 2)
            except Exception:
                pass
    return rates


# Bloco B: intervalos de código postal do distrito de Aveiro (aviso, não bloqueio).
# Fonte: CTT — o distrito de Aveiro cobre 3700-3899 (parte central+sul) e
# 4500-4550 (Espinho, Sta Maria da Feira, S. João da Madeira, zona norte).
# NOTA: se detectares moradas legítimas fora destes prefixos, ajusta aqui.
_AVEIRO_DISTRICT_POSTCODE_PREFIXES = (
    "37", "38",       # Aveiro cidade e envolvente (Águeda, Anadia, Ílhavo, ...)
    "45",             # Norte do distrito (Espinho, Feira, SJM, Arouca, Vale de Cambra, ...)
)


def _postcode_in_aveiro_district(code: str) -> bool:
    """Bloco B: verificação por prefixo — devolve True se o CP parece pertencer
    ao distrito. Conservador de propósito (só marca como fora em casos claros)."""
    clean = re.sub(r"[^0-9]", "", code or "")
    if len(clean) < 4:
        return True  # não temos dados suficientes — não avisamos
    return clean[:2] in _AVEIRO_DISTRICT_POSTCODE_PREFIXES


# Bloco C: validação do NIF português (algoritmo oficial do dígito de controlo)
# Primeiros dígitos válidos, oficiais:
#   1, 2, 3 → pessoas singulares
#   5       → pessoas coletivas (empresas)
#   6       → administração pública
#   8       → empresário em nome individual
#   9       → outras pessoas coletivas (associações, etc.)
# 4 e 7 NÃO são atribuídos como 1.º dígito de NIFs regulares em Portugal.
_VALID_NIF_FIRST_DIGITS = {"1", "2", "3", "5", "6", "8", "9"}


def validate_pt_nif(nif: str) -> bool:
    """Devolve True se o NIF for válido segundo o algoritmo oficial:
    1) exatamente 9 dígitos numéricos
    2) o 1.º dígito ∈ {1,2,3,5,6,8,9}
    3) checksum mód 11 sobre os primeiros 8 dígitos com pesos 9..2
       - se resto < 2  ⇒ dígito de controlo esperado = 0
       - caso contrário ⇒ dígito de controlo esperado = 11 − resto
    """
    n = re.sub(r"[^0-9]", "", nif or "")
    if len(n) != 9:
        return False
    if n[0] not in _VALID_NIF_FIRST_DIGITS:
        return False
    weights = [9, 8, 7, 6, 5, 4, 3, 2]
    total = sum(int(n[i]) * weights[i] for i in range(8))
    remainder = total % 11
    check = 0 if remainder < 2 else 11 - remainder
    return check == int(n[8])


# ------------ Shipping rates endpoints (Bloco B) ------------

@api.get("/shipping/zones")
async def public_shipping_zones():
    """Público: lista dos 19 concelhos + preço. Usado pelo dropdown do checkout."""
    rates = await _get_shipping_rates()
    return {
        "concelhos": [{"name": c, "rate": rates[c]} for c in AVEIRO_CONCELHOS],
    }


@api.get("/admin/shipping-rates")
async def admin_get_shipping_rates(admin: dict = Depends(require_admin)):
    return {"rates": await _get_shipping_rates(), "concelhos": AVEIRO_CONCELHOS}


class ShippingRatesIn(BaseModel):
    rates: Dict[str, float]


@api.put("/admin/shipping-rates")
async def admin_put_shipping_rates(payload: ShippingRatesIn, admin: dict = Depends(require_manager)):
    """Bloco B: substitui todos os custos de entrega por concelho. Aceita apenas
    concelhos que estejam na lista canónica AVEIRO_CONCELHOS."""
    clean: Dict[str, float] = {}
    for k, v in (payload.rates or {}).items():
        if k not in AVEIRO_CONCELHOS:
            continue  # ignora concelhos não canónicos, não dá erro para ser tolerante
        try:
            price = round(float(v), 2)
        except Exception:
            price = 0.0
        if price < 0:
            price = 0.0
        clean[k] = price
    await db.settings.update_one(
        {"id": "global"},
        {"$set": {"shipping_rates": clean, "id": "global"}},
        upsert=True,
    )
    await log_action(admin["id"], "update", "shipping_rates", "global", {"count": len(clean)})
    return {"rates": await _get_shipping_rates()}


# ------------ Legal pages (Bloco D) ------------

LEGAL_SLUGS = ("privacidade", "termos", "cookies", "ral")
LEGAL_TITLES = {
    "privacidade": "Política de Privacidade",
    "termos": "Termos e Condições",
    "cookies": "Política de Cookies",
    "ral": "Resolução Alternativa de Litígios",
}

# Tags permitidas no HTML do editor rico (Bloco D2). Qualquer <script>, on*, style
# malicioso é removido. Bleach lida com todo o clean-up.
_LEGAL_ALLOWED_TAGS = [
    "h1", "h2", "h3", "h4", "h5", "h6",
    "p", "br", "hr",
    "strong", "b", "em", "i", "u", "s", "del", "ins",
    "ul", "ol", "li",
    "blockquote", "code", "pre",
    "a", "span", "div",
]
_LEGAL_ALLOWED_ATTRS = {
    "a": ["href", "title", "target", "rel"],
    "*": [],
}


def _sanitize_legal_html(html: str) -> str:
    """Bloco D: sanitização defensiva. Remove <script>, event handlers e
    atributos perigosos. Ainda que o editor seja admin (confiável), evita
    danos acidentais e blinda contra XSS caso um dia haja um sub-admin."""
    import bleach
    return bleach.clean(
        html or "",
        tags=_LEGAL_ALLOWED_TAGS,
        attributes=_LEGAL_ALLOWED_ATTRS,
        strip=True,
    )


class LegalPageIn(BaseModel):
    content_html: str


@api.get("/legal/{slug}")
async def public_legal_page(slug: str):
    """Público: usado pelas páginas /legal/{slug} do site."""
    if slug not in LEGAL_SLUGS:
        raise HTTPException(404, "Página desconhecida")
    doc = await db.legal_pages.find_one({"slug": slug}, {"_id": 0}) or {}
    return {
        "slug": slug,
        "title": LEGAL_TITLES[slug],
        "content_html": doc.get("content_html") or "",
        "updated_at": doc.get("updated_at"),
    }


@api.get("/admin/legal")
async def admin_legal_list(admin: dict = Depends(require_admin)):
    """Devolve as 3 páginas legais (mesmo que ainda não tenham conteúdo)."""
    docs = {d["slug"]: d async for d in db.legal_pages.find({}, {"_id": 0})}
    return {
        "pages": [
            {
                "slug": s,
                "title": LEGAL_TITLES[s],
                "content_html": docs.get(s, {}).get("content_html", ""),
                "updated_at": docs.get(s, {}).get("updated_at"),
            }
            for s in LEGAL_SLUGS
        ]
    }


@api.put("/admin/legal/{slug}")
async def admin_legal_save(slug: str, payload: LegalPageIn, admin: dict = Depends(require_manager)):
    if slug not in LEGAL_SLUGS:
        raise HTTPException(404, "Página desconhecida")
    clean = _sanitize_legal_html(payload.content_html)
    await db.legal_pages.update_one(
        {"slug": slug},
        {"$set": {
            "slug": slug,
            "content_html": clean,
            "updated_at": iso(now_utc()),
            "updated_by": admin["id"],
        }},
        upsert=True,
    )
    await log_action(admin["id"], "update", "legal_page", slug, {"len": len(clean)})
    return {"ok": True, "slug": slug, "content_html": clean}


# ---------- Orders ----------
@api.post("/orders")
async def create_order(payload: OrderCreateIn):
    summary = await _compute_cart(payload.items, payload.promo_code, payload.bags_qty or 0)
    if not summary["lines"]:
        raise HTTPException(400, "Carrinho vazio ou livros indisponíveis")

    if payload.terms_accepted is not True:
        raise HTTPException(400, "É necessário aceitar os Termos e Condições para concluir a encomenda.")

    lamination_requested = any(bool(line.get("lamination")) for line in summary["lines"])
    if lamination_requested and payload.lamination_early_start_ack is not True:
        raise HTTPException(400, "Confirme a autorização para início antecipado do serviço de plastificação.")

    # Bloco C: se pediu fatura com NIF, valida NIF PT + exige nome fiscal
    if payload.wants_invoice:
        if not validate_pt_nif(payload.nif or ""):
            raise HTTPException(400, "NIF inválido, verifique.")
        if not (payload.fiscal_name or "").strip():
            raise HTTPException(400, "Nome fiscal em falta.")

    shipping_cost = 0.0
    delivery_concelho = (payload.delivery_concelho or "").strip() or None
    if payload.delivery_method == "hand_delivery":
        # Bloco B: concelho obrigatório para entrega em mão
        if not delivery_concelho:
            raise HTTPException(400, "Indique o concelho de entrega (distrito de Aveiro).")
        if delivery_concelho not in AVEIRO_CONCELHOS:
            raise HTTPException(400, "Concelho fora do distrito de Aveiro. Não fazemos entregas em mão fora do distrito.")
        rates = await _get_shipping_rates()
        shipping_cost = float(rates.get(delivery_concelho, 0.0))
        # Aviso não-bloqueante: se o código postal estiver fora do distrito
        # (aceita-se na mesma — pode haver exceções, ver spec Bloco B3).
    elif payload.delivery_method == "shipping":
        if not payload.address or not payload.postal_code:
            raise HTTPException(400, "Indique a morada e o código postal para envio.")
        settings = await db.settings.find_one({"id": "global"}, {"_id": 0}) or {}
        shipping_cost = float(settings.get("shipping_flat_rate", SHIPPING_FLAT_RATE))
    else:
        raise HTTPException(400, "Método de entrega inválido. Use 'hand_delivery' ou 'shipping'.")

    # Reserve stock atomically before committing the order (prevents overselling).
    stock_adjustments = await _reserve_stock(summary["lines"])

    total_with_shipping = round(summary["total"] + shipping_cost, 2)
    order_no = await _new_unique_order_no()
    access_token = _new_order_access_token()
    access_token_hash = _hash_order_access_token(access_token)
    access_token_expires_at = _order_access_expires_at()

    customer_match = await db.users.find_one(
        {"email": payload.customer_email.lower(), "role": "customer"}, {"_id": 0, "password_hash": 0}
    )

    terms_page = await db.legal_pages.find_one({"slug": "termos"}, {"_id": 0, "updated_at": 1}) or {}
    privacy_page = await db.legal_pages.find_one({"slug": "privacidade"}, {"_id": 0, "updated_at": 1}) or {}
    terms_version = terms_page.get("updated_at") or "unknown_terms_version"
    privacy_notice_version = privacy_page.get("updated_at") or "unknown_privacy_version"
    now_iso = iso(now_utc())
    lamination_ack_at = now_iso if lamination_requested and payload.lamination_early_start_ack else None

    doc = {
        "id": gen_id(),
        "order_no": order_no,
        "items": summary["lines"],
        "totals": {
            "subtotal_manuals": summary["subtotal_manuals"],
            "subtotal_workbooks": summary["subtotal_workbooks"],
            "discount_workbooks": summary["discount_workbooks"],
            "lamination_total": summary["lamination_total"],
            "bags_total": summary["bags_total"],
            "shipping_cost": round(shipping_cost, 2),
            "total": total_with_shipping,
        },
        "promo": summary.get("promo"),
        "customer": {
            "id": customer_match["id"] if customer_match else None,
            "name": payload.customer_name,
            "email": payload.customer_email.lower(),
            "phone": payload.customer_phone,
            # Bloco C: dados de faturação (guardados sempre, mesmo quando não pediu fatura)
            "wants_invoice": bool(payload.wants_invoice),
            "nif": (payload.nif or "").strip() if payload.wants_invoice else None,
            "fiscal_name": (payload.fiscal_name or "").strip() if payload.wants_invoice else None,
        },
        "delivery": {
            "method": payload.delivery_method,
            "concelho": delivery_concelho,
            "address": payload.address,
            "postal_code": payload.postal_code,
        },
        "notes": payload.notes,
        "bags_qty": max(0, int(payload.bags_qty or 0)),
        "status": "pending_payment",
        "payment_status": "pending",
        "payment_provider": "ifthenpay_mocked",
        "invoice_status": "not_issued",
        "stock_adjustments": stock_adjustments,
        "terms_accepted": True,
        "terms_accepted_at": now_iso,
        "terms_version": terms_version,
        "privacy_notice_version": privacy_notice_version,
        "lamination_early_start_ack": bool(lamination_requested and payload.lamination_early_start_ack),
        "lamination_ack_at": lamination_ack_at,
        "lamination_ack_version": 1 if lamination_ack_at else None,
        "access_token_hash": access_token_hash,
        "access_token_expires_at": iso(access_token_expires_at),
        "created_at": now_iso,
    }
    await db.orders.insert_one(doc)

    if summary.get("promo"):
        await db.partners.update_one({"promo_code": summary["promo"]["code"]}, {"$inc": {"usage_count": 1}})

    # MOCKED payment + invoice hook
    logger.info(f"[MOCKED IFTHENPAY] Order {order_no} created, total={total_with_shipping}€")
    return JSONResponse(
        content={
            "order": _order_confirmation_payload(doc),
            "access_token": access_token,
        },
        headers={"Cache-Control": "no-store"},
    )

@api.post("/orders/track")
async def track_order(payload: OrderTrackIn, request: Request):
    normalized_order_no = _normalize_order_no(payload.order_no)
    normalized_email = payload.email.strip().lower()
    origin_hint = _request_origin_hint(request)

    # Primary bucket: order_no + request origin. Varying the email must not
    # create fresh buckets for repeated brute-force attempts.
    await _check_order_access_rate_limit(
        scope="order_tracking",
        key_parts=[normalized_order_no, origin_hint],
        max_attempts=ORDER_TRACK_LIMIT_MAX_ATTEMPTS,
        window_seconds=ORDER_TRACK_LIMIT_WINDOW_SECONDS,
    )
    # Secondary guardrail: global attempts per origin across all order numbers.
    await _check_order_access_rate_limit(
        scope="order_tracking_origin",
        key_parts=[origin_hint],
        max_attempts=ORDER_TRACK_GLOBAL_LIMIT_MAX_ATTEMPTS,
        window_seconds=ORDER_TRACK_GLOBAL_LIMIT_WINDOW_SECONDS,
    )

    o = await db.orders.find_one({"order_no": normalized_order_no}, {"_id": 0})
    if not o:
        raise _public_order_track_error()

    order_email = str((o.get("customer") or {}).get("email") or "").strip().lower()
    if not order_email or not hmac.compare_digest(order_email, normalized_email):
        raise _public_order_track_error()

    return JSONResponse(content=_order_tracking_payload(o), headers={"Cache-Control": "no-store"})


@api.get("/orders/{order_no}")
async def get_order(order_no: str, request: Request):
    normalized_order_no = _normalize_order_no(order_no)
    origin_hint = _request_origin_hint(request)
    await _check_order_access_rate_limit(
        scope="order_confirmation",
        key_parts=[normalized_order_no, origin_hint],
        max_attempts=ORDER_CONFIRM_LIMIT_MAX_ATTEMPTS,
        window_seconds=ORDER_CONFIRM_LIMIT_WINDOW_SECONDS,
    )

    o = await db.orders.find_one({"order_no": normalized_order_no}, {"_id": 0})
    if not o:
        raise _public_order_access_error()

    token = request.headers.get(ORDER_ACCESS_TOKEN_HEADER)
    if not _is_valid_order_access_token(o, token):
        raise _public_order_access_error()

    return JSONResponse(content=_order_confirmation_payload(o), headers={"Cache-Control": "no-store"})

@api.get("/admin/orders")
async def admin_list_orders(admin: dict = Depends(require_admin), status: Optional[str] = None, archived: str = "false"):
    filt = {}
    if archived.lower() in ("true", "1", "yes"):
        filt["archived"] = True
    else:
        filt["archived"] = {"$ne": True}
    if status:
        filt["status"] = status
    return await db.orders.find(filt, {"_id": 0, "access_token_hash": 0}).sort("created_at", -1).to_list(500)

@api.put("/admin/orders/{order_no}/status")
async def admin_update_order(order_no: str, status: str = Form(...), admin: dict = Depends(require_admin)):
    order = await db.orders.find_one({"order_no": order_no})
    if not order:
        raise HTTPException(404, "Encomenda não encontrada")
    was_cancelled = (order.get("status") or "").lower().startswith("cancel")
    will_cancel = status.lower().startswith("cancel")
    if will_cancel and not was_cancelled:
        await _restore_stock(order.get("stock_adjustments", []))

    await db.orders.update_one({"order_no": order_no}, {"$set": {"status": status, "updated_at": iso(now_utc())}})
    await log_action(admin["id"], "update_status", "order", order_no, {"status": status})
    if status == "paid":
        logger.info(f"[MOCKED INVOICEXPRESS] Fatura-Recibo gerada para {order_no}")
        await db.orders.update_one({"order_no": order_no}, {"$set": {"invoice_status": "issued", "payment_status": "paid"}})
    return {"ok": True}

class ArchiveIn(BaseModel):
    ids: List[str]

@api.post("/admin/orders/archive")
async def admin_archive_orders(payload: ArchiveIn, admin: dict = Depends(require_admin)):
    """Archive multiple orders by order_no."""
    if not payload.ids:
        return {"archived": 0}
    result = await db.orders.update_many({"order_no": {"$in": payload.ids}}, {"$set": {"archived": True, "updated_at": iso(now_utc())}})
    await log_action(admin["id"], "archive", "orders", None, {"count": result.modified_count})
    return {"archived": result.modified_count}

@api.post("/admin/orders/unarchive")
async def admin_unarchive_orders(payload: ArchiveIn, admin: dict = Depends(require_admin)):
    """Unarchive multiple orders by order_no."""
    if not payload.ids:
        return {"unarchived": 0}
    result = await db.orders.update_many({"order_no": {"$in": payload.ids}}, {"$set": {"archived": False, "updated_at": iso(now_utc())}})
    await log_action(admin["id"], "unarchive", "orders", None, {"count": result.modified_count})
    return {"unarchived": result.modified_count}

# ---------- Vouchers ----------
async def _check_voucher_rate_limit(identifier: str):
    """Simple sliding-window throttle: max 8 voucher submissions per
    identifier (email or IP) per hour. Protects the public, unauthenticated
    endpoint from spam."""
    window_start = iso(now_utc() - timedelta(hours=1))
    count = await db.voucher_submissions.count_documents({"identifier": identifier, "at": {"$gte": window_start}})
    if count >= 8:
        raise HTTPException(429, "Demasiadas submissões de voucher. Tente novamente mais tarde ou contacte-nos.")
    await db.voucher_submissions.insert_one({"identifier": identifier, "at": iso(now_utc())})

@api.post("/vouchers")
async def submit_voucher(payload: VoucherSubmitIn, request: Request, user: Optional[dict] = Depends(get_current_user_optional)):
    """Voucher submission via code (no PDF). For PDF use POST /vouchers/upload."""
    identifier = (user["email"] if user else request.client.host) or "anon"
    await _check_voucher_rate_limit(identifier)
    code_clean = (payload.code or "").upper().strip() or None
    if code_clean and not re.match(r"^ALN\d{24}$", code_clean):
        raise HTTPException(400, "O código tem de ter o formato ALN seguido de 24 dígitos.")
    if not code_clean and not payload.pdf_url:
        raise HTTPException(400, "Indique um código ALN válido ou anexe o PDF do voucher.")
    if payload.name is not None and not payload.name.strip():
        raise HTTPException(400, "Nome é obrigatório.")
    if payload.contact is not None and not payload.contact.strip():
        raise HTTPException(400, "Contacto é obrigatório.")
    if payload.manuals is not None and not payload.manuals.strip():
        raise HTTPException(400, "Indique os manuais pretendidos.")
    doc = {
        "id": gen_id(),
        "name": (payload.name or "").strip() or None,
        "contact": (payload.contact or "").strip() or None,
        "manuals": (payload.manuals or "").strip() or None,
        "wants_workbooks": bool(payload.wants_workbooks),
        "workbook_details": (payload.workbook_details or "").strip() if payload.wants_workbooks else None,
        "wants_lamination": bool(payload.wants_lamination),
        "lamination_details": (payload.lamination_details or "").strip() if payload.wants_lamination else None,
        "code": code_clean,
        "pdf_url": payload.pdf_url,
        "pdf_storage_path": None,
        "notes": payload.notes,
        "status": "Pendente",
        "customer_id": user["id"] if user else None,
        "order_id": None,
        "created_at": iso(now_utc()),
    }
    await db.vouchers.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.post("/vouchers/upload")
async def upload_voucher(
    request: Request,
    file: UploadFile = File(...),
    code: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    name: Optional[str] = Form(None),
    contact: Optional[str] = Form(None),
    manuals: Optional[str] = Form(None),
    wants_workbooks: Optional[str] = Form("false"),
    workbook_details: Optional[str] = Form(None),
    wants_lamination: Optional[str] = Form("false"),
    lamination_details: Optional[str] = Form(None),
    user: Optional[dict] = Depends(get_current_user_optional),
):
    """Real, private PDF upload for a MEGA voucher. File is validated,
    renamed to a server-generated UUID, and stored OUTSIDE any publicly
    served directory. Readable only via GET /admin/vouchers/{id}/pdf."""
    identifier = (user["email"] if user else request.client.host) or "anon"
    await _check_voucher_rate_limit(identifier)

    if not name or not name.strip():
        raise HTTPException(400, "Nome é obrigatório.")
    if not contact or not contact.strip():
        raise HTTPException(400, "Contacto é obrigatório.")
    if not manuals or not manuals.strip():
        raise HTTPException(400, "Indique os manuais pretendidos.")

    code_clean = (code or "").upper().strip() or None
    if code_clean and not re.match(r"^ALN\d{24}$", code_clean):
        raise HTTPException(400, "O código tem de ter o formato ALN seguido de 24 dígitos.")

    content = await file.read()
    if len(content) > VOUCHER_MAX_BYTES:
        raise HTTPException(400, f"Ficheiro demasiado grande (máx. {VOUCHER_MAX_BYTES // (1024*1024)}MB).")
    if not content.startswith(b"%PDF"):
        raise HTTPException(400, "O ficheiro tem de ser um PDF válido.")

    stored_name = f"{gen_id()}.pdf"
    with open(VOUCHERS_DIR / stored_name, "wb") as f:
        f.write(content)

    doc = {
        "id": gen_id(),
        "name": name.strip(),
        "contact": contact.strip(),
        "manuals": manuals.strip(),
        "wants_workbooks": str(wants_workbooks).lower() in ("true", "1", "on", "yes"),
        "workbook_details": (workbook_details or "").strip() if str(wants_workbooks).lower() in ("true", "1", "on", "yes") else None,
        "wants_lamination": str(wants_lamination).lower() in ("true", "1", "on", "yes"),
        "lamination_details": (lamination_details or "").strip() if str(wants_lamination).lower() in ("true", "1", "on", "yes") else None,
        "code": code_clean,
        "pdf_url": None,
        "pdf_storage_path": stored_name,
        "notes": notes,
        "status": "Pendente",
        "customer_id": user["id"] if user else None,
        "order_id": None,
        "created_at": iso(now_utc()),
    }
    await db.vouchers.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.get("/admin/vouchers")
async def admin_vouchers(admin: dict = Depends(require_admin), status: Optional[str] = None, archived: str = "false"):
    filt = {}
    if archived.lower() in ("true", "1", "yes"):
        filt["archived"] = True
    else:
        filt["archived"] = {"$ne": True}
    if status:
        filt["status"] = status
    return await db.vouchers.find(filt, {"_id": 0}).sort("created_at", -1).to_list(500)

@api.get("/admin/vouchers/{vid}/pdf")
async def admin_voucher_pdf(vid: str, admin: dict = Depends(require_admin)):
    """The only way to read a voucher PDF — requires a valid admin/staff JWT.
    There is no public or guessable URL for these files."""
    voucher = await db.vouchers.find_one({"id": vid}, {"_id": 0})
    if not voucher or not voucher.get("pdf_storage_path"):
        raise HTTPException(404, "PDF não encontrado para este voucher")
    path = VOUCHERS_DIR / voucher["pdf_storage_path"]
    if not path.exists():
        raise HTTPException(404, "Ficheiro já não existe (pode ter sido removido por política de retenção)")
    return FileResponse(path, media_type="application/pdf", filename=f"voucher-{vid}.pdf")

@api.put("/admin/vouchers/{vid}/status")
async def admin_update_voucher(vid: str, status: str = Form(...), admin: dict = Depends(require_admin)):
    if status not in ("Pending", "Pendente", "Validated", "Used", "Rejected", "Em processamento", "Concluído"):
        raise HTTPException(400, "Estado inválido")
    await db.vouchers.update_one({"id": vid}, {"$set": {"status": status, "updated_at": iso(now_utc())}})
    await log_action(admin["id"], "update_status", "voucher", vid, {"status": status})
    voucher = await db.vouchers.find_one({"id": vid}, {"_id": 0})
    if voucher and voucher.get("customer_id"):
        cust = await db.users.find_one({"id": voucher["customer_id"]}, {"_id": 0})
        if cust:
            logger.info(f"[MOCKED EMAIL] Voucher {vid} -> {status} to {cust.get('email')}")
    return {"ok": True}

@api.put("/admin/vouchers/{vid}/note")
async def admin_update_voucher_note(vid: str, note: str = Form(""), admin: dict = Depends(require_admin)):
    """Update the internal admin note attached to a voucher (e.g. amount due
    for workbooks/lamination). Stored on the `notes` field of the voucher
    document. Empty string clears the note."""
    existing = await db.vouchers.find_one({"id": vid}, {"_id": 0, "id": 1})
    if not existing:
        raise HTTPException(404, "Voucher não encontrado")
    clean = (note or "").strip()
    await db.vouchers.update_one(
        {"id": vid},
        {"$set": {"notes": clean or None, "updated_at": iso(now_utc())}},
    )
    await log_action(admin["id"], "update_note", "voucher", vid, {"len": len(clean)})
    return {"ok": True, "notes": clean or None}

@api.post("/admin/vouchers/archive")
async def admin_archive_vouchers(payload: ArchiveIn, admin: dict = Depends(require_admin)):
    """Archive multiple vouchers by id."""
    if not payload.ids:
        return {"archived": 0}
    result = await db.vouchers.update_many({"id": {"$in": payload.ids}}, {"$set": {"archived": True, "updated_at": iso(now_utc())}})
    await log_action(admin["id"], "archive", "vouchers", None, {"count": result.modified_count})
    return {"archived": result.modified_count}

@api.post("/admin/vouchers/unarchive")
async def admin_unarchive_vouchers(payload: ArchiveIn, admin: dict = Depends(require_admin)):
    """Unarchive multiple vouchers by id."""
    if not payload.ids:
        return {"unarchived": 0}
    result = await db.vouchers.update_many({"id": {"$in": payload.ids}}, {"$set": {"archived": False, "updated_at": iso(now_utc())}})
    await log_action(admin["id"], "unarchive", "vouchers", None, {"count": result.modified_count})
    return {"unarchived": result.modified_count}

async def _purge_old_voucher_pdfs():
    """There is no cron infra in this environment, so this sweep runs
    once per backend startup instead of daily. For a long-running production
    deployment, wire this into a real scheduler (e.g. APScheduler/Celery
    beat) so it also runs between restarts."""
    cutoff = iso(now_utc() - timedelta(days=VOUCHER_RETENTION_DAYS))
    async for v in db.vouchers.find({"pdf_storage_path": {"$ne": None}, "created_at": {"$lt": cutoff}}, {"_id": 0}):
        path = VOUCHERS_DIR / v["pdf_storage_path"]
        try:
            if path.exists():
                path.unlink()
        except Exception:
            pass
        await db.vouchers.update_one({"id": v["id"]}, {"$set": {"pdf_storage_path": None, "pdf_purged_at": iso(now_utc())}})

# ---------- Wishlist ----------
@api.get("/wishlist")
async def get_wishlist(user: dict = Depends(get_current_user)):
    items = await db.wishlist.find({"user_id": user["id"]}, {"_id": 0}).to_list(500)
    isbns = [i["isbn13"] for i in items]
    books = await db.books.find({"isbn13": {"$in": isbns}}, {"_id": 0}).to_list(500)
    return books

@api.post("/wishlist")
async def add_wishlist(payload: WishlistIn, user: dict = Depends(get_current_user)):
    isbn = strip_isbn(payload.isbn13)
    if not await db.wishlist.find_one({"user_id": user["id"], "isbn13": isbn}):
        await db.wishlist.insert_one({"id": gen_id(), "user_id": user["id"], "isbn13": isbn, "created_at": iso(now_utc())})
    return {"ok": True}

@api.delete("/wishlist/{isbn13}")
async def remove_wishlist(isbn13: str, user: dict = Depends(get_current_user)):
    await db.wishlist.delete_one({"user_id": user["id"], "isbn13": strip_isbn(isbn13)})
    return {"ok": True}

# ---------- Admin: users, settings, logs ----------
@api.get("/admin/dashboard")
async def admin_dashboard(admin: dict = Depends(require_admin)):
    total_books = await db.books.count_documents({})
    total_schools = await db.schools.count_documents({})
    total_orders = await db.orders.count_documents({"archived": {"$ne": True}})
    pending_vouchers = await db.vouchers.count_documents({"status": "Pending", "archived": {"$ne": True}})
    anomalies = await db.books.count_documents({"$or": [{"price": 0}, {"price": {"$lte": 0}}]})
    recent_orders = await db.orders.find({"archived": {"$ne": True}}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    return {
        "total_books": total_books,
        "total_schools": total_schools,
        "total_orders": total_orders,
        "pending_vouchers": pending_vouchers,
        "anomalies": anomalies,
        "recent_orders": recent_orders,
    }

# ---------- Customers (distinct from admin/staff Users) ----------
@api.get("/admin/customers")
async def admin_list_customers(admin: dict = Depends(require_admin), q: Optional[str] = None):
    filt: Dict[str, Any] = {"role": "customer"}
    if q:
        regex = {"$regex": re.escape(q), "$options": "i"}
        filt["$or"] = [{"name": regex}, {"email": regex}]
    return await db.users.find(filt, {"_id": 0, "password_hash": 0}).sort("created_at", -1).to_list(1000)

@api.get("/admin/customers/{cid}")
async def admin_customer_detail(cid: str, admin: dict = Depends(require_admin)):
    cust = await db.users.find_one({"id": cid, "role": "customer"}, {"_id": 0, "password_hash": 0})
    if not cust:
        raise HTTPException(404, "Cliente não encontrado")
    orders = await db.orders.find({"customer.id": cid}, {"_id": 0}).sort("created_at", -1).to_list(200)
    vouchers = await db.vouchers.find({"customer_id": cid}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return {"customer": cust, "orders": orders, "vouchers": vouchers}

@api.put("/admin/customers/{cid}/block")
async def admin_block_customer(cid: str, blocked: bool = Form(...), admin: dict = Depends(require_manager)):
    """Block/unblock — manager-tier only. Staff must not be able to do this."""
    res = await db.users.update_one({"id": cid, "role": "customer"}, {"$set": {"is_blocked": blocked}})
    if res.matched_count == 0:
        raise HTTPException(404, "Cliente não encontrado")
    await log_action(admin["id"], "block" if blocked else "unblock", "customer", cid)
    return {"ok": True}

@api.delete("/admin/customers/{cid}")
async def admin_delete_customer(cid: str, admin: dict = Depends(require_manager)):
    """Delete — manager-tier only. Staff must not be able to do this."""
    res = await db.users.delete_one({"id": cid, "role": "customer"})
    if res.deleted_count == 0:
        raise HTTPException(404, "Cliente não encontrado")
    await log_action(admin["id"], "delete", "customer", cid)
    return {"ok": True}

# ---------- Reports (financial — manager-tier only, NOT staff) ----------
@api.get("/admin/reports")
async def admin_reports(admin: dict = Depends(require_manager), days: int = 365):
    since = iso(now_utc() - timedelta(days=days))
    paid_filter = {"payment_status": "paid", "created_at": {"$gte": since}}

    revenue_total = 0.0
    monthly: Dict[str, float] = {}
    book_sales: Dict[str, Dict[str, Any]] = {}
    async for o in db.orders.find(paid_filter, {"_id": 0}):
        total = float((o.get("totals") or {}).get("total", 0) or 0)
        revenue_total += total
        month_key = (o.get("created_at") or "")[:7]
        monthly[month_key] = round(monthly.get(month_key, 0.0) + total, 2)
        for it in o.get("items", []):
            isbn = it.get("isbn13")
            if not isbn:
                continue
            rec = book_sales.setdefault(isbn, {"isbn13": isbn, "title": it.get("title", ""), "qty": 0, "revenue": 0.0})
            rec["qty"] += it.get("qty", 0)
            rec["revenue"] = round(rec["revenue"] + it.get("line_total", 0), 2)

    bestsellers = sorted(book_sales.values(), key=lambda r: r["qty"], reverse=True)[:10]
    total_orders_paid = await db.orders.count_documents(paid_filter)
    pending_payment = await db.orders.count_documents({"status": "pending_payment"})

    return {
        "revenue_total": round(revenue_total, 2),
        "monthly_revenue": dict(sorted(monthly.items())),
        "bestsellers": bestsellers,
        "paid_orders": total_orders_paid,
        "pending_payment_orders": pending_payment,
        "period_days": days,
    }

@api.get("/admin/users")
async def admin_users(admin: dict = Depends(require_super_admin)):
    return await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)

@api.get("/admin/activity-logs")
async def admin_logs(admin: dict = Depends(require_admin), limit: int = 200):
    return await db.activity_logs.find({}, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(limit)

@api.get("/admin/settings")
async def get_settings(admin: dict = Depends(require_manager)):
    s = await db.settings.find_one({"id": "global"}, {"_id": 0}) or {
        "id": "global",
        "lamination_price": LAMINATION_PRICE,
        "shipping_flat_rate": SHIPPING_FLAT_RATE,
        "aveiro_postcodes": _AVEIRO_CONCELHO_POSTCODES,
    }
    return s

@api.put("/admin/settings")
async def update_settings(payload: SettingIn, admin: dict = Depends(require_super_admin)):
    update = {k: v for k, v in payload.model_dump().items() if v is not None}
    await db.settings.update_one({"id": "global"}, {"$set": {**update, "id": "global"}}, upsert=True)
    await log_action(admin["id"], "update", "settings", "global", update)
    return await db.settings.find_one({"id": "global"}, {"_id": 0})

# ---------- Seeding ----------
async def ensure_indexes():
    await db.users.create_index("email", unique=True)
    # Bloco A: livros podem ter ISBN OU Código PE (mutuamente exclusivos, ou ambos).
    # Migração segura do índice antigo "isbn13_1 unique" para um partial-unique
    # (só único quando isbn13 é não-vazio, permitindo múltiplos livros sem ISBN).
    try:
        existing = await db.books.index_information()
        # Se o índice antigo existe (unique estrito, sem partialFilter), removê-lo
        old = existing.get("isbn13_1")
        if old is not None and old.get("unique") and "partialFilterExpression" not in old:
            await db.books.drop_index("isbn13_1")
    except Exception as e:
        logger.warning(f"[indexes] Não foi possível inspecionar/limpar índice isbn13_1: {e}")
    # Partial-unique: só livros com isbn13 não-vazio são forçados a ser únicos.
    # Livros sem ISBN (só código PE) podem coexistir com isbn13="".
    await db.books.create_index(
        "isbn13",
        unique=True,
        partialFilterExpression={"isbn13": {"$type": "string", "$gt": ""}},
        name="isbn13_unique_when_present",
    )
    # pe_code único quando presente
    await db.books.create_index(
        "pe_code",
        unique=True,
        partialFilterExpression={"pe_code": {"$type": "string", "$gt": ""}},
        name="pe_code_unique_when_present",
    )
    # slug único quando presente (para URLs SEO-friendly)
    await db.books.create_index(
        "slug",
        unique=True,
        partialFilterExpression={"slug": {"$type": "string", "$gt": ""}},
        name="slug_unique_when_present",
    )
    await db.school_books.create_index([("school_id", 1), ("isbn13", 1), ("grade_level", 1)])
    await db.login_attempts.create_index("identifier")
    await db.password_reset_tokens.create_index("expires_at")
    await db.categories.create_index("name", unique=True)
    await db.partners.create_index("promo_code", unique=True)
    await db.voucher_submissions.create_index([("identifier", 1), ("at", 1)])
    await db.orders.create_index("order_no", name="order_no_lookup_idx")

    # Create a unique index only when historical data has no duplicates.
    duplicate_cursor = db.orders.aggregate([
        {"$group": {"_id": "$order_no", "count": {"$sum": 1}}},
        {"$match": {"_id": {"$ne": None}, "count": {"$gt": 1}}},
        {"$limit": 1},
    ])
    duplicates = await duplicate_cursor.to_list(length=1)
    if duplicates:
        logger.warning("[orders] Duplicados históricos em order_no detetados; índice unique não foi criado.")
    else:
        try:
            await db.orders.create_index("order_no", unique=True, name="order_no_unique_if_safe")
        except Exception as e:
            logger.warning(f"[orders] Não foi possível criar índice unique para order_no: {e}")

    await db.order_access_rate_limits.create_index(
        [("scope", 1), ("identifier_hash", 1), ("created_at_dt", 1)],
        name="order_access_rate_lookup_idx",
    )
    await db.order_access_rate_limits.create_index("expires_at", expireAfterSeconds=0, name="order_access_rate_ttl")

async def seed_admins():
    admins = [
        {"email": os.environ["SUPER_ADMIN_EMAIL"], "password": os.environ["SUPER_ADMIN_PASSWORD"], "role": "super_admin", "name": "Jetro Manança"},
        {"email": os.environ["ADMIN_EMAIL"], "password": os.environ["ADMIN_PASSWORD"], "role": "admin", "name": "F. Tendinha"},
    ]
    for a in admins:
        existing = await db.users.find_one({"email": a["email"].lower()})
        if not existing:
            await db.users.insert_one({
                "id": gen_id(),
                "email": a["email"].lower(),
                "name": a["name"],
                "role": a["role"],
                "password_hash": hash_password(a["password"]),
                "must_change_password": False,
                "created_at": iso(now_utc()),
            })
            logger.info(f"Seeded {a['role']}: {a['email']}")
        elif not verify_password(a["password"], existing["password_hash"]):
            await db.users.update_one(
                {"email": a["email"].lower()},
                {"$set": {"password_hash": hash_password(a["password"]), "role": a["role"], "name": a["name"]}},
            )

def _grade_to_label(s: str) -> str:
    s = str(s or "").strip()
    digits = re.search(r"(\d+)", s)
    return f"{digits.group(1)}.º Ano" if digits else s

def _parse_grade_list(s: str) -> List[str]:
    s = str(s or "").strip()
    if not s:
        return []
    tokens = re.split(r"[;,]", s)
    parsed: List[str] = []
    for token in tokens:
        token = str(token or "").strip()
        if not token:
            continue
        if re.search(r"[-–—]", token):
            parts = [p.strip() for p in re.split(r"[-–—]", token) if p.strip()]
            if len(parts) == 2:
                start = _grade_to_label(parts[0])
                end = _grade_to_label(parts[1])
                if start in _GRADES_ALL and end in _GRADES_ALL:
                    start_idx = _GRADES_ALL.index(start)
                    end_idx = _GRADES_ALL.index(end)
                    if start_idx <= end_idx:
                        parsed.extend(_GRADES_ALL[start_idx:end_idx + 1])
                    else:
                        parsed.extend(_GRADES_ALL[end_idx:start_idx + 1])
                    continue
        normalized = _grade_to_label(token)
        if normalized in _GRADES_ALL:
            parsed.append(normalized)
    # Preserve order and remove duplicates.
    unique: List[str] = []
    for grade in parsed:
        if grade not in unique:
            unique.append(grade)
    return unique

_GRADES_ALL = ["1.º Ano", "2.º Ano", "3.º Ano", "4.º Ano", "5.º Ano", "6.º Ano",
               "7.º Ano", "8.º Ano", "9.º Ano", "10.º Ano", "11.º Ano", "12.º Ano",
               "Profissional", "Secundário Profissional"]
_GRADES_EB1 = ["1.º Ano", "2.º Ano", "3.º Ano", "4.º Ano"]
_GRADES_EB23 = ["5.º Ano", "6.º Ano", "7.º Ano", "8.º Ano", "9.º Ano"]
_GRADES_SEC = ["10.º Ano", "11.º Ano", "12.º Ano"]
_GRADES_EB_SEC = _GRADES_EB23 + _GRADES_SEC

# Postcode prefixes for the concelho (município) de Aveiro ONLY — NOT the
# wider Aveiro district. The district has 19 different municípios (Ílhavo,
# Vagos, Estarreja, Ovar, Espinho, Santa Maria da Feira, etc.) which must
# NOT be included here, since "entrega em mão" is concelho-scoped per spec.
# Verified against CTT/codigo-postal.pt: 3800-38xx and 3810-38xx cover the
# Aveiro city freguesias (Esgueira, Aradas, Glória, Vera Cruz, Santa Joana,
# Cacia, Eixo e Eirol, Oliveirinha, S. Bernardo, S. Jacinto, Requeixo).
# This is an editable STARTING default (see /admin/settings) — Francisco/
# Jetro should fine-tune the exact prefix list to match the real delivery
# radius they're willing to cover on foot/bike within the concelho.
_AVEIRO_CONCELHO_POSTCODES = [
    "3800", "3801", "3802", "3803", "3804",
    "3810", "3811", "3812", "3813", "3814",
]

def _cover_url(isbn: str) -> str:
    # Books start with NO cover URL. Real covers are fetched on demand from
    # multiple sources via the "Procurar capas" action (see /admin/books/enrich-covers).
    # Returning "" means the storefront shows its themed placeholder until a
    # genuine cover is found, instead of a broken Open Library image.
    return ""

async def seed_demo_data():
    if await db.municipalities.count_documents({}) > 0:
        return
    # Aveiro district — 19 municipalities
    muns_data = [
        "Águeda", "Albergaria-a-Velha", "Anadia", "Arouca", "Aveiro",
        "Castelo de Paiva", "Espinho", "Estarreja", "Ílhavo", "Mealhada",
        "Murtosa", "Oliveira de Azeméis", "Oliveira do Bairro", "Ovar",
        "Santa Maria da Feira", "São João da Madeira", "Sever do Vouga",
        "Vagos", "Vale de Cambra",
    ]
    mun_ids = {}
    for m in muns_data:
        mid = gen_id()
        await db.municipalities.insert_one({"id": mid, "name": m})
        mun_ids[m] = mid

    # Schools — multiple per municipality covering different cycles
    # (name, municipality, grades_taught)
    schools_data = [
        # Aveiro (concelho de referência — todas as escolas reais conhecidas)
        ("Agrupamento de Escolas de Aveiro", "Aveiro", _GRADES_ALL),
        ("Escola Secundária José Estêvão", "Aveiro", _GRADES_SEC),
        ("Escola Secundária Homem Cristo", "Aveiro", _GRADES_SEC),
        ("Escola Secundária Dr. Mário Sacramento", "Aveiro", _GRADES_SEC),
        ("EB 2,3 João Afonso de Aveiro", "Aveiro", _GRADES_EB23),
        ("EB 2,3 de Esgueira", "Aveiro", _GRADES_EB23),
        ("EB 1 de Aveiro – Glória", "Aveiro", _GRADES_EB1),
        ("EB 1 de Aveiro – Vera Cruz", "Aveiro", _GRADES_EB1),
        ("EB 1 de Esgueira", "Aveiro", _GRADES_EB1),
        ("EB 1 de Cacia", "Aveiro", _GRADES_EB1),
        # Ílhavo
        ("Escola Secundária Dr. João Carlos Celestino Gomes", "Ílhavo", _GRADES_SEC),
        ("EB 2,3 Gafanha da Nazaré", "Ílhavo", _GRADES_EB23),
        ("EB 2,3 José Ferreira Pinto Basto", "Ílhavo", _GRADES_EB23),
        ("EB 1 da Gafanha da Encarnação", "Ílhavo", _GRADES_EB1),
        ("EB 1 da Vista Alegre", "Ílhavo", _GRADES_EB1),
        # Vagos
        ("Escola Secundária de Vagos", "Vagos", _GRADES_SEC),
        ("EB 2,3 de Vagos", "Vagos", _GRADES_EB23),
        ("EB 1 de Vagos", "Vagos", _GRADES_EB1),
        # Águeda
        ("Escola Secundária Marques de Castilho", "Águeda", _GRADES_SEC),
        ("Escola Secundária Adolfo Portela", "Águeda", _GRADES_SEC),
        ("EB 2,3 Fernando Caldeira", "Águeda", _GRADES_EB23),
        ("EB 1 de Águeda", "Águeda", _GRADES_EB1),
        # Oliveira do Bairro
        ("Escola Secundária de Oliveira do Bairro", "Oliveira do Bairro", _GRADES_SEC),
        ("EB 2,3 Acácio de Azevedo", "Oliveira do Bairro", _GRADES_EB23),
        ("EB 1 de Oliveira do Bairro", "Oliveira do Bairro", _GRADES_EB1),
        # Ovar
        ("Escola Secundária Júlio Dinis", "Ovar", _GRADES_SEC),
        ("EB 2,3 Florbela Espanca", "Ovar", _GRADES_EB23),
        ("EB 1 de Ovar", "Ovar", _GRADES_EB1),
        # Estarreja
        ("Escola Secundária de Estarreja", "Estarreja", _GRADES_SEC),
        ("EB 2,3 Padre Donaciano de Abreu Freire", "Estarreja", _GRADES_EB23),
        # Albergaria-a-Velha
        ("Escola Secundária de Albergaria-a-Velha", "Albergaria-a-Velha", _GRADES_SEC),
        ("EB 2,3 de Albergaria-a-Velha", "Albergaria-a-Velha", _GRADES_EB23),
        # Anadia
        ("Escola Secundária de Anadia", "Anadia", _GRADES_SEC),
        ("EB 2,3 de Anadia", "Anadia", _GRADES_EB23),
        # Arouca
        ("Escola Secundária de Arouca", "Arouca", _GRADES_SEC),
        ("EB 2,3 de Arouca", "Arouca", _GRADES_EB23),
        # Espinho
        ("Escola Secundária Dr. Manuel Laranjeira", "Espinho", _GRADES_SEC),
        ("EB 2,3 Sá Couto", "Espinho", _GRADES_EB23),
        # Murtosa
        ("EB 2,3 / Secundária da Murtosa", "Murtosa", _GRADES_EB_SEC),
        # Mealhada
        ("Escola Secundária da Mealhada", "Mealhada", _GRADES_SEC),
        ("EB 2,3 da Mealhada", "Mealhada", _GRADES_EB23),
        # Oliveira de Azeméis
        ("Escola Secundária Dr. Ferreira da Silva", "Oliveira de Azeméis", _GRADES_SEC),
        ("EB 2,3 de Cucujães", "Oliveira de Azeméis", _GRADES_EB23),
        # São João da Madeira
        ("Escola Secundária Serafim Leite", "São João da Madeira", _GRADES_SEC),
        ("EB 2,3 de S. João da Madeira", "São João da Madeira", _GRADES_EB23),
        # Vale de Cambra
        ("Escola Secundária de Vale de Cambra", "Vale de Cambra", _GRADES_SEC),
        # Santa Maria da Feira
        ("Escola Secundária Coelho e Castro", "Santa Maria da Feira", _GRADES_SEC),
        ("EB 2,3 de Lobão", "Santa Maria da Feira", _GRADES_EB23),
        # Sever do Vouga
        ("Escola Secundária de Sever do Vouga", "Sever do Vouga", _GRADES_SEC),
        # Castelo de Paiva
        ("Escola Secundária de Castelo de Paiva", "Castelo de Paiva", _GRADES_SEC),
    ]
    school_ids = {}
    for entry in schools_data:
        name, mun, grades = entry
        sid = gen_id()
        await db.schools.insert_one({"id": sid, "name": name, "municipality_id": mun_ids[mun], "grades_taught": grades})
        school_ids[name] = sid

    # Import real catalog from seed Excel if present
    seed_xlsx = ROOT_DIR / "seed_catalog.xlsx"
    imported_isbn_grade = {}  # isbn -> grade label
    if seed_xlsx.exists():
        try:
            df = pd.read_excel(seed_xlsx)
            for _, row in df.iterrows():
                isbn = strip_isbn(str(row.get("ISBN", "")))
                if len(isbn) != 13:
                    continue
                try:
                    price = float(row.get("PVP") or 0)
                except Exception:
                    price = 0.0
                if price <= 0:
                    continue
                artigo = str(row.get("Artigo", "")).strip().lower()
                book_type = "Workbook" if "caderno" in artigo else "Manual"
                title = str(row.get("Título") or "").strip()
                subject = str(row.get("Disciplina") or "").strip()
                publisher = str(row.get("Editora") or "").strip()
                author = str(row.get("Autor(es)") or "").strip()
                grade = _grade_to_label(row.get("Ano"))
                cycle = str(row.get("Ciclo de Ensino") or "").strip()
                doc = {
                    "id": gen_id(),
                    "isbn13": isbn,
                    "title": title,
                    "author": author,
                    "publisher": publisher,
                    "subject": subject,
                    "year": 2025,
                    "price": round(price, 2),
                    "type": book_type,
                    "status": "Available",
                    "stock_qty": 15,
                    "synopsis": f"{title} — material escolar oficial para o {grade}, da editora {publisher}.",
                    "features": {"cycle": cycle, "grade": grade, "format": book_type},
                    "image_url": _cover_url(isbn),
                    "is_lamination_eligible": True,  # lamination applies to Manuals AND Workbooks alike
                    "created_at": iso(now_utc()),
                }
                try:
                    await db.books.insert_one(doc)
                    imported_isbn_grade[isbn] = grade
                except Exception:
                    pass
            logger.info(f"Imported {len(imported_isbn_grade)} books from seed_catalog.xlsx")
        except Exception as e:
            logger.warning(f"Could not import seed catalog: {e}")

    # Additional curated demo books (used only if Excel import is missing or small)
    books_seed = [] if imported_isbn_grade else [
        # Manuals
        {"isbn13": "9789897070945", "title": "Pasta Mágica 1 - Português", "author": "Angelina Rodrigues", "publisher": "Areal Editores", "year": 2024, "subject": "Português", "price": 24.50, "type": "Manual", "is_lamination_eligible": True, "synopsis": "Manual de Português do 1.º ano de escolaridade, alinhado com as aprendizagens essenciais.", "image_url": "https://images.unsplash.com/photo-1497633762265-9d179a990aa6?w=400&q=80"},
        {"isbn13": "9789897078811", "title": "Alfa Matemática 1.º Ano", "author": "Eva Lima", "publisher": "Porto Editora", "year": 2024, "subject": "Matemática", "price": 26.90, "type": "Manual", "is_lamination_eligible": True, "synopsis": "Manual de Matemática, com foco em raciocínio e resolução de problemas.", "image_url": "https://images.unsplash.com/photo-1543002588-bfa74002ed7e?w=400&q=80"},
        {"isbn13": "9789897078812", "title": "Estudo do Meio 1.º Ano", "author": "Ana Maria Pessoa", "publisher": "Porto Editora", "year": 2024, "subject": "Estudo do Meio", "price": 22.40, "type": "Manual", "is_lamination_eligible": True, "synopsis": "Descobrir o mundo através de atividades práticas.", "image_url": "https://images.unsplash.com/photo-1532012197267-da84d127e765?w=400&q=80"},
        {"isbn13": "9789897078813", "title": "Diálogos 5 - Português", "author": "Helena Margarida Vaz", "publisher": "Porto Editora", "year": 2024, "subject": "Português", "price": 32.10, "type": "Manual", "is_lamination_eligible": True, "synopsis": "Manual de Português para o 5.º ano.", "image_url": "https://images.unsplash.com/photo-1544947950-fa07a98d237f?w=400&q=80"},
        {"isbn13": "9789897078814", "title": "MSI 5 - Matemática", "author": "Maria Augusta Ferreira Neves", "publisher": "Porto Editora", "year": 2024, "subject": "Matemática", "price": 32.10, "type": "Manual", "is_lamination_eligible": True, "synopsis": "Matemática Sob Investigação - 5.º ano.", "image_url": "https://images.unsplash.com/photo-1635070041078-e363dbe005cb?w=400&q=80"},
        {"isbn13": "9789897078815", "title": "Português 9 - Mensagens", "author": "Lúcia Vidal Soares", "publisher": "Texto Editores", "year": 2024, "subject": "Português", "price": 34.80, "type": "Manual", "is_lamination_eligible": True, "synopsis": "Manual de Português para o 9.º ano.", "image_url": "https://images.unsplash.com/photo-1456513080510-7bf3a84b82f8?w=400&q=80"},
        {"isbn13": "9789897078816", "title": "Novo Espaço 11 - Matemática A", "author": "Belmiro Costa", "publisher": "Porto Editora", "year": 2024, "subject": "Matemática A", "price": 38.50, "type": "Manual", "is_lamination_eligible": True, "synopsis": "Manual de Matemática A para o 11.º ano.", "image_url": "https://images.unsplash.com/photo-1518744386442-2d48ac47a7eb?w=400&q=80"},
        # Workbooks
        {"isbn13": "9789897078901", "title": "Caderno de Fichas Pasta Mágica 1", "author": "Angelina Rodrigues", "publisher": "Areal Editores", "year": 2024, "subject": "Português", "price": 12.30, "type": "Workbook", "is_lamination_eligible": True, "synopsis": "Caderno de fichas de Português 1.º ano.", "image_url": "https://images.unsplash.com/photo-1455390582262-044cdead277a?w=400&q=80"},
        {"isbn13": "9789897078902", "title": "Caderno Alfa Matemática 1.º", "author": "Eva Lima", "publisher": "Porto Editora", "year": 2024, "subject": "Matemática", "price": 12.80, "type": "Workbook", "is_lamination_eligible": True, "synopsis": "Caderno de atividades de Matemática 1.º ano.", "image_url": "https://images.unsplash.com/photo-1513475382585-d06e58bcb0e0?w=400&q=80"},
        {"isbn13": "9789897078903", "title": "Caderno Diálogos 5", "author": "Helena Vaz", "publisher": "Porto Editora", "year": 2024, "subject": "Português", "price": 14.20, "type": "Workbook", "is_lamination_eligible": True, "synopsis": "Caderno de atividades de Português 5.º ano.", "image_url": "https://images.unsplash.com/photo-1497633762265-9d179a990aa6?w=400&q=80"},
        {"isbn13": "9789897078904", "title": "Caderno MSI 5", "author": "Maria Neves", "publisher": "Porto Editora", "year": 2024, "subject": "Matemática", "price": 14.50, "type": "Workbook", "is_lamination_eligible": True, "synopsis": "Caderno de fichas de Matemática 5.º ano.", "image_url": "https://images.unsplash.com/photo-1576094792933-2f29e7e5fe71?w=400&q=80"},
        {"isbn13": "9789897078905", "title": "Caderno Mensagens 9", "author": "Lúcia Soares", "publisher": "Texto Editores", "year": 2024, "subject": "Português", "price": 15.10, "type": "Workbook", "is_lamination_eligible": True, "synopsis": "Caderno de Português 9.º ano.", "image_url": "https://images.unsplash.com/photo-1491841550275-ad7854e35ca6?w=400&q=80"},
        {"isbn13": "9789897078906", "title": "Caderno Novo Espaço 11", "author": "Belmiro Costa", "publisher": "Porto Editora", "year": 2024, "subject": "Matemática A", "price": 16.40, "type": "Workbook", "is_lamination_eligible": True, "synopsis": "Caderno de Matemática A 11.º ano.", "image_url": "https://images.unsplash.com/photo-1503676260728-1c00da094a0b?w=400&q=80"},
        {"isbn13": "9789897078920", "title": "Inglês 5 - Step Up", "author": "Cristina Cunha", "publisher": "Porto Editora", "year": 2024, "subject": "Inglês", "price": 28.40, "type": "Manual", "is_lamination_eligible": True, "synopsis": "Manual de Inglês 5.º ano.", "image_url": "https://images.unsplash.com/photo-1546410531-bb4caa6b424d?w=400&q=80", "status": "PreOrder", "stock_qty": 0},
        {"isbn13": "9789897078921", "title": "Físico-Química 8 - Eureka!", "author": "Filomena Caldeira", "publisher": "Asa Editores", "year": 2024, "subject": "Físico-Química", "price": 30.20, "type": "Manual", "is_lamination_eligible": True, "synopsis": "Manual de FQ 8.º ano.", "image_url": "https://images.unsplash.com/photo-1554475901-4538ddfbccc2?w=400&q=80", "status": "Unavailable", "stock_qty": 0},
    ]
    for b in books_seed:
        b.setdefault("status", "Available")
        b.setdefault("stock_qty", 25)
        b.setdefault("synopsis", "")
        b.setdefault("features", {"format": "Manual escolar", "pages": 192})
        b["id"] = gen_id()
        b["created_at"] = iso(now_utc())
        await db.books.insert_one(b)

    # Link books to schools: associate every imported book to schools that teach that grade
    for isbn, grade in imported_isbn_grade.items():
        async for s in db.schools.find({"grades_taught": grade}, {"_id": 0, "id": 1}):
            await db.school_books.insert_one({
                "id": gen_id(), "school_id": s["id"], "isbn13": isbn, "grade_level": grade,
            })

    # Partners (real Aveiro partners) — logos via UI Avatars (hotlink-friendly)
    partners_data = [
        {"name": "Academia do Beira-Mar", "logo_url": "https://ui-avatars.com/api/?name=Beira+Mar&background=000000&color=ffffff&size=240&font-size=0.45&bold=true", "description": "Os atletas da Academia do Beira-Mar beneficiam de um desconto exclusivo nos cadernos de fichas.", "promo_code": "BEIRAMAR5", "discount_value": 5.0, "order": 1, "active": True},
        {"name": "Academia Vista Alegre", "logo_url": "https://ui-avatars.com/api/?name=Vista+Alegre&background=1B4965&color=ffffff&size=240&font-size=0.40&bold=true", "description": "Os atletas da Academia Vista Alegre beneficiam de um desconto exclusivo nos cadernos de fichas.", "promo_code": "VISTAALEGRE5", "discount_value": 5.0, "order": 2, "active": True},
        {"name": "Iliabum Clube", "logo_url": "https://ui-avatars.com/api/?name=Iliabum+Clube&background=5A8F1E&color=ffffff&size=240&font-size=0.42&bold=true", "description": "Os atletas do Iliabum Clube beneficiam de um desconto exclusivo nos cadernos de fichas.", "promo_code": "ILIABUM5", "discount_value": 5.0, "order": 3, "active": True},
    ]
    for p in partners_data:
        p["id"] = gen_id()
        p.setdefault("usage_count", 0)
        p.setdefault("valid_from", None)
        p.setdefault("valid_until", None)
        p.setdefault("usage_limit", None)
        await db.partners.insert_one(p)

    # Settings
    await db.settings.insert_one({
        "id": "global",
        "lamination_price": LAMINATION_PRICE,
        "shipping_flat_rate": SHIPPING_FLAT_RATE,
        "aveiro_postcodes": _AVEIRO_CONCELHO_POSTCODES,
    })
    logger.info("Demo data seeded")

@app.on_event("startup")
async def startup():
    await ensure_indexes()
    await seed_admins()
    await seed_demo_data()
    await _purge_old_voucher_pdfs()

@app.on_event("shutdown")
async def shutdown():
    client.close()

# ---------- Health ----------
@api.get("/")
async def root():
    return {"message": "Tendinha do Saber API", "version": "2.0"}

# ---------- SEO endpoints (public) ----------
@api.get("/seo/tracking")
async def seo_tracking():
    s = await db.settings.find_one({"id": "global"}, {"_id": 0}) or {}
    return {
        "google_analytics_id": s.get("google_analytics_id") or "",
        "google_ads_id": s.get("google_ads_id") or "",
        "facebook_pixel_id": s.get("facebook_pixel_id") or "",
        "google_site_verification": s.get("google_site_verification") or "",
    }

@api.get("/seo/sitemap.xml")
async def seo_sitemap(request: Request):
    from fastapi.responses import Response
    s = await db.settings.find_one({"id": "global"}, {"_id": 0}) or {}
    base = s.get("site_url") or str(request.base_url).rstrip("/").replace("/api", "")
    # Static public pages
    static_paths = ["/", "/catalogo", "/parceiros", "/vouchers", "/como-funciona-voucher",
                    "/sobre", "/faq", "/contactos", "/seguir-encomenda",
                    "/legal/privacidade", "/legal/termos", "/legal/cookies", "/legal/ral"]
    today = now_utc().date().isoformat()
    urls = [f"<url><loc>{base}{p}</loc><lastmod>{today}</lastmod><changefreq>weekly</changefreq><priority>{'1.0' if p=='/' else '0.7'}</priority></url>" for p in static_paths]
    # All books
    async for b in db.books.find({"status": {"$ne": "Unavailable"}}, {"isbn13": 1, "_id": 0}):
        urls.append(f"<url><loc>{base}/livro/{b['isbn13']}</loc><lastmod>{today}</lastmod><changefreq>weekly</changefreq><priority>0.6</priority></url>")
    xml = f"""<?xml version=\"1.0\" encoding=\"UTF-8\"?>
<urlset xmlns=\"http://www.sitemaps.org/schemas/sitemap/0.9\">
{''.join(urls)}
</urlset>"""
    return Response(content=xml, media_type="application/xml")

# ---------- Register router ----------
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)
